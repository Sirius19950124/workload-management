# -*- coding: utf-8 -*-
"""
惠阳妇幼保健院康复科业务管理系统 - Excel导入导出
V1.0 - 2026-02-24
"""

import os
import io
from datetime import datetime, date, timedelta
from flask import Blueprint, request, jsonify, send_file, current_app
from sqlalchemy import extract, func
from app import db
from app.api.log_decorator import log_op
from app.models import (
    WorkloadTherapist, WorkloadTreatmentCategory,
    WorkloadTreatmentItem, WorkloadRecord
)

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

workload_excel_bp = Blueprint('workload_excel', __name__, url_prefix='/api/excel')


@workload_excel_bp.route('/import', methods=['POST'])
@log_op('import', 'Excel导入工作量记录')
def import_from_excel():
    """从Excel文件导入数据"""
    if not EXCEL_AVAILABLE:
        return jsonify({'success': False, 'error': 'Excel库未安装，请运行: pip install pandas openpyxl'}), 500

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '请上传Excel文件'}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': '请上传Excel文件 (.xlsx 或 .xls)'}), 400

    try:
        # 读取Excel文件
        xl = pd.ExcelFile(file)

        # 导入结果统计
        result = {
            'therapists_created': 0,
            'treatment_items_created': 0,
            'records_created': 0,
            'errors': []
        }

        # 1. 尝试导入数据源表（治疗项目和治疗师）
        source_sheets = [s for s in xl.sheet_names if '源' in s or 'source' in s.lower()]
        if source_sheets:
            df_source = pd.read_excel(file, sheet_name=source_sheets[0], header=None)

            # 自动检测列索引
            header_row = df_source.iloc[0].tolist()
            col_map = {}
            for idx, col_name in enumerate(header_row):
                col_str = str(col_name).strip() if pd.notna(col_name) else ''
                col_map[col_str] = idx

            # 查找各列的索引
            code_col = None
            category_col = None
            item_col = None
            weight_col = None

            for col_name, idx in col_map.items():
                if col_name in ['编号', '项目编号', '代码', 'code', 'Code']:
                    code_col = idx
                elif col_name in ['治疗类别', '类别', '分类', '项目类别', 'category', 'Category']:
                    category_col = idx
                elif col_name in ['治疗项目', '项目名称', '具体治疗项目名称', '名称', 'name', 'Name', 'item']:
                    item_col = idx
                elif col_name in ['权重', '权重系数', '综合权重系数', 'weight', 'Weight']:
                    weight_col = idx

            # 如果没有找到列，尝试按位置推断（兼容旧格式）
            if item_col is None:
                # 尝试查找包含"项目"的列
                for col_name, idx in col_map.items():
                    if '项目' in col_name or '名称' in col_name:
                        item_col = idx
                        break

            result['debug_info'] = {
                'detected_columns': col_map,
                'code_col': code_col,
                'category_col': category_col,
                'item_col': item_col,
                'weight_col': weight_col
            }

            for idx, row in df_source.iterrows():
                if idx == 0:  # 跳过标题行
                    continue

                # 根据检测到的列获取值
                if item_col is not None and len(row) > item_col and pd.notna(row[item_col]):
                    item_name = str(row[item_col]).strip()

                    # 跳过说明文字和非项目名称
                    skip_item_keywords = ['提示', '说明', '模板', '参考', '下拉', '自动计算', '治疗师', '总权重', '人次']
                    if any(kw in item_name for kw in skip_item_keywords):
                        continue
                    if len(item_name) > 20 or item_name.isdigit():
                        continue

                    # 获取类别
                    category_name = None
                    if category_col is not None and len(row) > category_col and pd.notna(row[category_col]):
                        category_name = str(row[category_col]).strip()

                    # 获取编号
                    code = None
                    if code_col is not None and len(row) > code_col and pd.notna(row[code_col]):
                        code = str(row[code_col]).strip()

                    # 获取权重
                    weight = 1.0
                    if weight_col is not None and len(row) > weight_col and pd.notna(row[weight_col]):
                        try:
                            weight = float(row[weight_col])
                        except:
                            weight = 1.0

                    # 创建或获取类别
                    category = None
                    if category_name:
                        category = WorkloadTreatmentCategory.query.filter_by(name=category_name).first()
                        if not category:
                            category = WorkloadTreatmentCategory(name=category_name)
                            db.session.add(category)
                            db.session.flush()

                    # 创建治疗项目（按名称或编号查重）
                    existing = None
                    if code:
                        existing = WorkloadTreatmentItem.query.filter_by(code=code).first()
                    if not existing:
                        existing = WorkloadTreatmentItem.query.filter_by(name=item_name).first()

                    if not existing:
                        item = WorkloadTreatmentItem(
                            code=code,
                            name=item_name,
                            category_id=category.id if category else None,
                            weight_coefficient=weight
                        )
                        db.session.add(item)
                        result['treatment_items_created'] += 1

            # 尝试解析治疗师（通常在数据源表右侧）
            # 需要跳过的非人名词汇（说明、表头等）
            skip_words = [
                '治疗师', '项目', '总权重', '人次', 'nan',
                '提示', '说明', '参考', '备注', '姓名', '工号',
                '治疗类别', '权重系数', '编号', '代码', '序号',
                '日期', '患者', '加权工作量', '治疗项目',
                '导入模板', '工作量记录', '下拉菜单', '自动计算',
            ]
            for col_idx in range(5, min(len(df_source.columns), 15)):
                # 先检查该列第一个值是否是'治疗师'表头
                if pd.notna(df_source.iloc[0, col_idx]) and str(df_source.iloc[0, col_idx]).strip() != '治疗师':
                    continue
                col_values = df_source.iloc[:, col_idx].dropna()
                for val in col_values:
                    if pd.notna(val) and isinstance(val, str) and len(val.strip()) > 0:
                        name = val.strip()
                        # 跳过说明/非人名词汇
                        if name in skip_words:
                            continue
                        # 跳过包含说明性文字的行
                        if any(kw in name for kw in ['提示', '说明', '模板', '参考', '下拉', '自动']):
                            continue
                        # 跳过纯数字（可能是编号）
                        if name.isdigit():
                            continue
                        # 跳过过长的文本（说明文字通常较长）
                        if len(name) > 10:
                            continue
                        if not WorkloadTherapist.query.filter_by(name=name).first():
                            therapist = WorkloadTherapist(name=name)
                            db.session.add(therapist)
                            result['therapists_created'] += 1

            db.session.commit()

        # 2. 导入各日期工作表的数据
        for sheet_name in xl.sheet_names:
            if sheet_name in ['数据源', '汇总看板', '月度汇总'] or '源' in sheet_name:
                continue

            try:
                df = pd.read_excel(file, sheet_name=sheet_name, header=None)

                # ---- 智能检测表头行和列映射 ----
                header_row = None
                col_map = {}  # 列名 -> 列索引
                for i in range(min(5, len(df))):
                    row_values = df.iloc[i].tolist()
                    row_str = str(row_values)
                    # 更严格的表头检测：必须同时包含"日期"和"治疗师"
                    has_date = any('日期' in str(v) for v in row_values if pd.notna(v))
                    has_therapist = any('治疗师' in str(v) for v in row_values if pd.notna(v))
                    if has_date and has_therapist:
                        header_row = i
                        # 构建列名->索引的映射
                        for col_idx, val in enumerate(row_values):
                            if pd.notna(val):
                                col_str = str(val).strip()
                                if col_str:
                                    col_map[col_str] = col_idx
                        result['debug_info']['sheet_{}'.format(sheet_name)] = {
                            'header_row': header_row,
                            'col_map': col_map
                        }
                        break

                if header_row is None:
                    result['errors'].append(f'工作表 {sheet_name}: 未找到有效表头(需包含"日期"和"治疗师")')
                    continue

                # 根据检测到的表头确定各列索引（支持中英文和常见别名）
                def _find_col(*names):
                    """在col_map中查找列索引，支持多个候选名称"""
                    for n in names:
                        for k, v in col_map.items():
                            if n.lower() in k.lower() or k.lower() in n.lower():
                                return v
                    return None

                date_col = _find_col('日期')
                therapist_col = _find_col('治疗师')
                patient_col = _find_col('患者', '病人', '患者ID', '姓名')
                item_col = _find_col('治疗项目', '项目', '具体治疗项目名称')
                weight_col = _find_col('权重', '权重系数', '综合权重系数', '权值')
                sessions_col = _find_col('人次', '次数', 'session')
                remark_col = _find_col('备注', '说明', 'remark')

                # 如果没检测到某些列，使用基于位置的回退策略
                # 兼容旧格式：日期(0) 治疗师(1) 患者(2) 项目(3) 权重(4) 人次(5) 备注(6/7)
                if date_col is None: date_col = 0
                if therapist_col is None: therapist_col = 1
                if patient_col is None: patient_col = 2
                if item_col is None: item_col = 3
                # 权重和人次的位置取决于是否有权重列
                if weight_col is not None and sessions_col is None:
                    # 有权重列但未识别人次列 → 人次通常在权重后面
                    sessions_col = weight_col + 1
                elif weight_col is None and sessions_col is None:
                    # 无权重列 → 人次在第5列（兼容旧6列格式）
                    sessions_col = 5

                for idx in range(header_row + 1, len(df)):
                    row = df.iloc[idx]

                    # ---- 日期 ----
                    record_date = None
                    if date_col is not None and date_col < len(row) and pd.notna(row[date_col]):
                        try:
                            val = row[date_col]
                            if isinstance(val, datetime):
                                record_date = val.date()
                            elif isinstance(val, date):
                                record_date = val
                            else:
                                parsed = pd.to_datetime(val)
                                record_date = parsed.date() if hasattr(parsed, 'date') else parsed
                        except Exception:
                            continue

                    if not record_date:
                        continue

                    # ---- 治疗师 ----
                    therapist_name = None
                    if therapist_col is not None and therapist_col < len(row) and pd.notna(row[therapist_col]):
                        therapist_name = str(row[therapist_col]).strip()
                    if not therapist_name:
                        continue

                    therapist = WorkloadTherapist.query.filter_by(name=therapist_name).first()
                    if not therapist:
                        therapist = WorkloadTherapist(name=therapist_name)
                        db.session.add(therapist)
                        db.session.flush()
                        result['therapists_created'] += 1

                    # ---- 患者 ----
                    patient_info = None
                    if patient_col is not None and patient_col < len(row) and pd.notna(row[patient_col]):
                        patient_info = str(row[patient_col]).strip()

                    # ---- 治疗项目 ----
                    item_name = None
                    if item_col is not None and item_col < len(row) and pd.notna(row[item_col]):
                        item_name = str(row[item_col]).strip()
                    if not item_name:
                        continue

                    treatment_item = WorkloadTreatmentItem.query.filter_by(name=item_name).first()
                    if not treatment_item:
                        item_weight = 1.0
                        if weight_col is not None and weight_col < len(row) and pd.notna(row[weight_col]):
                            try:
                                item_weight = float(row[weight_col])
                            except:
                                pass
                        treatment_item = WorkloadTreatmentItem(
                            name=item_name,
                            weight_coefficient=item_weight
                        )
                        db.session.add(treatment_item)
                        db.session.flush()
                        result['treatment_items_created'] += 1

                    # ---- 权重系数 ----
                    weight = treatment_item.weight_coefficient
                    if weight_col is not None and weight_col < len(row) and pd.notna(row[weight_col]):
                        try:
                            weight = float(row[weight_col])
                        except:
                            pass

                    # ---- 人次（关键修复！）----
                    sessions = 1
                    if sessions_col is not None and sessions_col < len(row) and pd.notna(row[sessions_col]):
                        try:
                            sessions = int(float(row[sessions_col]))
                            if sessions < 0:
                                sessions = 1
                        except:
                            sessions = 1

                    # ---- 备注 ----
                    remark = None
                    if remark_col is not None and remark_col < len(row) and pd.notna(row[remark_col]):
                        remark = str(row[remark_col]).strip()

                    existing = WorkloadRecord.query.filter_by(
                        record_date=record_date,
                        therapist_id=therapist.id,
                        treatment_item_id=treatment_item.id,
                        patient_info=patient_info
                    ).first()

                    if not existing:
                        record = WorkloadRecord(
                            record_date=record_date,
                            therapist_id=therapist.id,
                            patient_info=patient_info,
                            treatment_item_id=treatment_item.id,
                            weight_coefficient=weight,
                            session_count=sessions,
                            weighted_workload=round(weight * sessions, 2),
                            remark=remark
                        )
                        db.session.add(record)
                        result['records_created'] += 1

                db.session.commit()

            except Exception as e:
                result['errors'].append(f'工作表 {sheet_name} 导入失败: {str(e)}')
                db.session.rollback()

        return jsonify({
            'success': True,
            'message': '导入完成',
            'data': result
        })

    except Exception as e:
        return jsonify({'success': False, 'error': f'导入失败: {str(e)}'}), 500


@workload_excel_bp.route('/export', methods=['GET'])
def export_to_excel():
    """导出数据到Excel文件"""
    if not EXCEL_AVAILABLE:
        return jsonify({'success': False, 'error': 'Excel库未安装'}), 500

    year_month = request.args.get('month', date.today().strftime('%Y-%m'))

    try:
        year, month = map(int, year_month.split('-'))
    except ValueError:
        return jsonify({'success': False, 'error': '月份格式错误'}), 400

    try:
        records = WorkloadRecord.query.filter(
            extract('year', WorkloadRecord.record_date) == year,
            extract('month', WorkloadRecord.record_date) == month
        ).order_by(WorkloadRecord.record_date, WorkloadRecord.therapist_id).all()

        therapists = WorkloadTherapist.query.filter_by(is_active=True).order_by(WorkloadTherapist.sort_order).all()
        items = WorkloadTreatmentItem.query.filter_by(is_active=True).all()

        wb = openpyxl.Workbook()

        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font_white = Font(bold=True, color='FFFFFF', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')

        # 1. 创建数据源表
        ws_source = wb.active
        ws_source.title = '数据源'

        ws_source['A1'] = '序号'
        ws_source['B1'] = '治疗类别'
        ws_source['C1'] = '具体治疗项目名称'
        ws_source['D1'] = '综合权重系数'

        for i, item in enumerate(items, 1):
            ws_source.cell(row=i+1, column=1, value=i)
            ws_source.cell(row=i+1, column=2, value=item.category_rel.name if item.category_rel else '其他')
            ws_source.cell(row=i+1, column=3, value=item.name)
            ws_source.cell(row=i+1, column=4, value=item.weight_coefficient)

        ws_source['F1'] = '治疗师'
        for i, t in enumerate(therapists, 1):
            ws_source.cell(row=i+1, column=6, value=t.name)

        # 2. 按日期创建工作表
        record_dates = sorted(set(r.record_date for r in records))

        for record_date in record_dates:
            day = record_date.day
            ws = wb.create_sheet(title=str(day))

            headers = ['日期', '治疗师', '患者ID/姓名', '治疗项目', '权重系数', '人次', '加权工作量', '备注']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_align

            ws.cell(row=1, column=1, value='提示：请从下拉菜单中选择治疗项目，权重和工作量自动计算')

            day_records = [r for r in records if r.record_date == record_date]
            for row_idx, record in enumerate(day_records, 3):
                ws.cell(row=row_idx, column=1, value=record.record_date)
                ws.cell(row=row_idx, column=2, value=record.therapist_rel.name if record.therapist_rel else '')
                ws.cell(row=row_idx, column=3, value=record.patient_info or '')
                ws.cell(row=row_idx, column=4, value=record.treatment_item_rel.name if record.treatment_item_rel else '')
                ws.cell(row=row_idx, column=5, value=record.weight_coefficient)
                ws.cell(row=row_idx, column=6, value=record.session_count)
                ws.cell(row=row_idx, column=7, value=record.weighted_workload)
                ws.cell(row=row_idx, column=8, value=record.remark or '')

            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 8
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 15

        # 3. 创建汇总看板
        ws_dashboard = wb.create_sheet(title='汇总看板')

        today = date.today()
        today_records = [r for r in records if r.record_date == today]

        ws_dashboard['A1'] = '今日日期'
        ws_dashboard['B1'] = today
        ws_dashboard['A2'] = '今日总人次'
        ws_dashboard['B2'] = sum(r.session_count for r in today_records)
        ws_dashboard['A3'] = '今日总加权工作量'
        ws_dashboard['B3'] = round(sum(r.weighted_workload for r in today_records), 2)

        ws_dashboard['E1'] = '月份'
        ws_dashboard['F1'] = '治疗师'
        ws_dashboard['G1'] = '月总工作量'

        therapist_monthly = {}
        for r in records:
            tid = r.therapist_id
            if tid not in therapist_monthly:
                therapist_monthly[tid] = {'name': r.therapist_rel.name, 'workload': 0}
            therapist_monthly[tid]['workload'] += r.weighted_workload

        row = 2
        for tid, stats in sorted(therapist_monthly.items(), key=lambda x: x[1]['workload'], reverse=True):
            ws_dashboard.cell(row=row, column=5, value=year_month)
            ws_dashboard.cell(row=row, column=6, value=stats['name'])
            ws_dashboard.cell(row=row, column=7, value=round(stats['workload'], 2))
            row += 1

        # 4. 创建月度汇总表
        ws_monthly = wb.create_sheet(title='月度汇总')

        ws_monthly.cell(row=1, column=1, value='治疗师')
        ws_monthly.cell(row=1, column=2, value='日期')

        import calendar
        _, days_in_month = calendar.monthrange(year, month)

        for day in range(1, days_in_month + 1):
            ws_monthly.cell(row=1, column=day + 2, value=day)

        for row_idx, therapist in enumerate(therapists, 2):
            ws_monthly.cell(row=row_idx, column=1, value=therapist.name)

            t_records = [r for r in records if r.therapist_id == therapist.id]

            daily_workload = {}
            for r in t_records:
                day = r.record_date.day
                if day not in daily_workload:
                    daily_workload[day] = 0
                daily_workload[day] += r.weighted_workload

            for day, workload in daily_workload.items():
                ws_monthly.cell(row=row_idx, column=day + 2, value=round(workload, 2))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'工作量统计_{year_month}.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'success': False, 'error': f'导出失败: {str(e)}'}), 500


@workload_excel_bp.route('/template', methods=['GET'])
def download_template():
    """下载Excel模板"""
    if not EXCEL_AVAILABLE:
        return jsonify({'success': False, 'error': 'Excel库未安装'}), 500

    try:
        wb = openpyxl.Workbook()

        ws_source = wb.active
        ws_source.title = '数据源'

        headers = ['编号', '治疗类别', '治疗项目', '权重系数']
        for col, header in enumerate(headers, 1):
            ws_source.cell(row=1, column=col, value=header)

        sample_items = [
            ('T001', '仪器治疗', '盆底电', 1.83),
            ('T002', '仪器治疗', '盆底磁', 0.83),
            ('T003', '手法治疗', '腰椎手法', 10.08),
            ('T004', '手法治疗', '颈椎手法', 5.92),
        ]

        for row_idx, (seq, category, name, weight) in enumerate(sample_items, 2):
            ws_source.cell(row=row_idx, column=1, value=seq)
            ws_source.cell(row=row_idx, column=2, value=category)
            ws_source.cell(row=row_idx, column=3, value=name)
            ws_source.cell(row=row_idx, column=4, value=weight)

        ws_source.cell(row=1, column=6, value='治疗师')
        sample_therapists = ['李德裕', '陈宇凡', '白芮', '张婷']
        for row_idx, name in enumerate(sample_therapists, 2):
            ws_source.cell(row=row_idx, column=6, value=name)

        ws_day = wb.create_sheet(title='1')

        ws_day.cell(row=1, column=1, value='提示：请从下拉菜单中选择治疗项目，权重和工作量自动计算')

        day_headers = ['日期', '治疗师', '患者ID/姓名', '治疗项目', '权重系数', '人次', '加权工作量', '备注']
        for col, header in enumerate(day_headers, 1):
            ws_day.cell(row=2, column=col, value=header)

        ws_day.cell(row=3, column=1, value=date.today())
        ws_day.cell(row=3, column=2, value='李德裕')
        ws_day.cell(row=3, column=3, value='张三')
        ws_day.cell(row=3, column=4, value='腰椎手法')
        ws_day.cell(row=3, column=5, value=10.08)
        ws_day.cell(row=3, column=6, value=1)
        ws_day.cell(row=3, column=7, value=10.08)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='工作量登记模板.xlsx'
        )

    except Exception as e:
        return jsonify({'success': False, 'error': f'生成模板失败: {str(e)}'}), 500
