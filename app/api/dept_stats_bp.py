# -*- coding: utf-8 -*-
"""
科室工作量统计 & 转介统计 - 完全重写版
三类数据独立：门诊、病房、儿童医院
基础项目库：录入时下拉选择，不用手敲
"""

from datetime import datetime, date, timedelta
import re
import json
from flask import Blueprint, request, jsonify, send_file
from sqlalchemy import text, inspect
from app import db
from app.api.log_decorator import log_op
import openpyxl
from io import BytesIO

dept_stats_bp = Blueprint('dept_stats', __name__, url_prefix='/api/dept-stats')
referral_bp = Blueprint('referral', __name__, url_prefix='/api/referral')

# ============================================================================
# 建表
# ============================================================================

TABLES_SQL = [
    # --- 基础项目库 ---
    """CREATE TABLE IF NOT EXISTS dept_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        item_type VARCHAR(20) NOT NULL,
        item_name VARCHAR(100) NOT NULL,
        sub_category VARCHAR(100) DEFAULT '',
        is_active BOOLEAN DEFAULT 1,
        sort_order INTEGER DEFAULT 0,
        UNIQUE(item_type, item_name, sub_category)
    )""",
    # --- 转介医生库 ---
    """CREATE TABLE IF NOT EXISTS referral_doctors (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        department VARCHAR(100) NOT NULL,
        doctor_name VARCHAR(50) NOT NULL,
        is_active BOOLEAN DEFAULT 1,
        UNIQUE(department, doctor_name)
    )""",
    # --- 转介指标库 ---
    """CREATE TABLE IF NOT EXISTS referral_metrics (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        department VARCHAR(100) DEFAULT '',
        metric_name VARCHAR(100) NOT NULL,
        is_active BOOLEAN DEFAULT 1,
        UNIQUE(department, metric_name)
    )""",
    # --- 门诊月度数据（独立表） ---
    """CREATE TABLE IF NOT EXISTS outpatient_monthly (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        item_name VARCHAR(100) NOT NULL,
        year INTEGER NOT NULL,
        month INTEGER NOT NULL,
        session_count INTEGER DEFAULT 0,
        amount FLOAT DEFAULT 0,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(item_name, year, month)
    )""",
    # --- 病房月度数据（独立表） ---
    """CREATE TABLE IF NOT EXISTS ward_monthly (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ward_area VARCHAR(100) NOT NULL,
        item_name VARCHAR(100) NOT NULL,
        year INTEGER NOT NULL,
        month INTEGER NOT NULL,
        session_count INTEGER DEFAULT 0,
        amount FLOAT DEFAULT 0,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(ward_area, item_name, year, month)
    )""",
    # --- 儿童医院月度汇总（独立表） ---
    """CREATE TABLE IF NOT EXISTS children_monthly (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        item_name VARCHAR(100) NOT NULL,
        year INTEGER NOT NULL,
        month INTEGER NOT NULL,
        session_count INTEGER DEFAULT 0,
        amount FLOAT DEFAULT 0,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(item_name, year, month)
    )""",
    # --- 儿童医院每日明细（独立表） ---
    """CREATE TABLE IF NOT EXISTS children_daily (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        record_date DATE NOT NULL,
        item_name VARCHAR(100) NOT NULL,
        session_count INTEGER DEFAULT 0,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(record_date, item_name)
    )""",
    """CREATE TABLE IF NOT EXISTS children_doctor_monthly (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        doctor_name VARCHAR(100) NOT NULL,
        year INTEGER NOT NULL,
        month INTEGER NOT NULL,
        session_count INTEGER DEFAULT 0,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(doctor_name, year, month)
    )""",
    # --- 转介月度数据（独立表） ---
    """CREATE TABLE IF NOT EXISTS referral_monthly (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        department VARCHAR(100) NOT NULL,
        doctor_name VARCHAR(50) NOT NULL,
        year INTEGER NOT NULL,
        month INTEGER NOT NULL,
        metric_name VARCHAR(100) NOT NULL,
        metric_count INTEGER DEFAULT 0,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(department, doctor_name, year, month, metric_name)
    )""",
    # --- 病房每日数据（透视表录入） ---
    """CREATE TABLE IF NOT EXISTS ward_daily (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        record_date DATE NOT NULL,
        ward_area VARCHAR(100) NOT NULL DEFAULT '',
        item_name VARCHAR(100) NOT NULL,
        session_count INTEGER DEFAULT 0,
        amount FLOAT DEFAULT 0,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(record_date, ward_area, item_name)
    )""",
    # --- 病房项目单价配置 ---
    """CREATE TABLE IF NOT EXISTS ward_item_prices (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ward_area VARCHAR(100) NOT NULL DEFAULT '',
        item_name VARCHAR(100) NOT NULL,
        unit_price FLOAT NOT NULL DEFAULT 0,
        is_active BOOLEAN DEFAULT 1,
        sort_order INTEGER DEFAULT 0,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(ward_area, item_name)
    )""",
    # --- 病房每日数据快照（回滚用） ---
    """CREATE TABLE IF NOT EXISTS ward_daily_snapshots (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ward_area VARCHAR(100) NOT NULL,
        snapshot_time DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
        trigger_type VARCHAR(50) NOT NULL DEFAULT 'auto_save',
        data_json TEXT NOT NULL,
        record_count INTEGER NOT NULL DEFAULT 0
    )""",
    # --- 通用数据快照（各模块回滚用） ---
    """CREATE TABLE IF NOT EXISTS data_snapshots (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        module VARCHAR(50) NOT NULL,
        scope_key VARCHAR(200) NOT NULL,
        table_name VARCHAR(100) NOT NULL,
        key_columns TEXT NOT NULL,
        snapshot_time DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
        trigger_type VARCHAR(50) NOT NULL DEFAULT 'auto_save',
        data_json TEXT NOT NULL,
        record_count INTEGER NOT NULL DEFAULT 0
    )""",
]


def ensure_tables():
    for sql in TABLES_SQL:
        db.session.execute(text(sql))
    db.session.commit()
    # Schema migration: add columns that may be missing from older databases
    _migrate_schema()


def _migrate_schema():
    """Add missing columns to existing tables (SQLite doesn't auto-add new cols)"""
    migrations = [
        ('children_monthly', 'amount', 'FLOAT DEFAULT 0'),
        ('outpatient_monthly', 'amount', 'FLOAT DEFAULT 0'),
        ('ward_monthly', 'amount', 'FLOAT DEFAULT 0'),
        ('referral_monthly', 'metric_count', 'INTEGER DEFAULT 0'),
        ('dept_items', 'unit_price', 'REAL DEFAULT 0'),
    ]
    inspector = inspect(db.engine)
    for table, col, col_type in migrations:
        if table in inspector.get_table_names():
            existing_cols = [c['name'] for c in inspector.get_columns(table)]
            if col not in existing_cols:
                try:
                    db.session.execute(text(f'ALTER TABLE {table} ADD COLUMN {col} {col_type}'))
                    db.session.commit()
                except Exception:
                    db.session.rollback()


# ============================================================================
# 一、基础项目库管理
# ============================================================================

@dept_stats_bp.route('/items', methods=['GET'])
def get_items():
    """获取项目列表（按类型）"""
    ensure_tables()
    item_type = request.args.get('type', 'outpatient')
    sub = request.args.get('sub', '')
    
    query = "SELECT id, item_name, sub_category, is_active FROM dept_items WHERE item_type = :type"
    params = {'type': item_type}
    if sub:
        query += " AND sub_category = :sub"
        params['sub'] = sub
    query += " ORDER BY sort_order, item_name"
    
    rows = db.session.execute(text(query), params).fetchall()
    return jsonify({'success': True, 'data': [
        {'id': r[0], 'name': r[1], 'sub': r[2], 'active': r[3]} for r in rows
    ]})


@dept_stats_bp.route('/items', methods=['POST'])
@log_op('create', '新增科室项目')
def add_item():
    """添加项目"""
    ensure_tables()
    data = request.get_json()
    item_type = data.get('type', 'outpatient')
    item_name = data.get('name', '').strip()
    sub_category = data.get('sub', '').strip()
    
    if not item_name:
        return jsonify({'success': False, 'error': '项目名称不能为空'}), 400
    
    try:
        db.session.execute(text("""
            INSERT OR IGNORE INTO dept_items (item_type, item_name, sub_category)
            VALUES (:type, :name, :sub)
        """), {'type': item_type, 'name': item_name, 'sub': sub_category})
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@dept_stats_bp.route('/items/<int:item_id>', methods=['DELETE'])
@log_op('delete', '删除科室项目')
def delete_item(item_id):
    """删除项目"""
    ensure_tables()
    db.session.execute(text("DELETE FROM dept_items WHERE id = :id"), {'id': item_id})
    db.session.commit()
    return jsonify({'success': True})


@dept_stats_bp.route('/items/batch', methods=['POST'])
def batch_add_items():
    """批量添加项目"""
    ensure_tables()
    data = request.get_json()
    item_type = data.get('type', 'outpatient')
    names = data.get('names', [])
    sub = data.get('sub', '').strip()
    
    count = 0
    for name in names:
        name = name.strip()
        if not name:
            continue
        db.session.execute(text("""
            INSERT OR IGNORE INTO dept_items (item_type, item_name, sub_category)
            VALUES (:type, :name, :sub)
        """), {'type': item_type, 'name': name, 'sub': sub})
        count += 1
    
    db.session.commit()
    return jsonify({'success': True, 'data': {'count': count}})


@dept_stats_bp.route('/items/import-excel', methods=['POST'])
@log_op('import', '科室统计导入Excel')
def import_items_excel():
    """从Excel批量导入项目到项目库+月度数据"""
    ensure_tables()
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '请上传文件'}), 400

    item_type = request.form.get('type', 'outpatient')
    sub = request.form.get('sub', '').strip()
    year = request.form.get('year', '').strip()
    file = request.files['file']

    # 解析months参数（前端JSON字符串）
    months_raw = request.form.get('months', '[]')
    try:
        import json as _json
        months_list = _json.loads(months_raw) if months_raw else []
    except Exception:
        months_list = []

    if not year:
        year = str(datetime.now().year)

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
    except Exception as e:
        return jsonify({'success': False, 'error': f'无法读取Excel文件: {str(e)}'}), 400

    count = 0
    monthly_count = 0
    errors = []
    sort_idx = 0
    skip_sheet_keywords = ('说明', '填写说明', '使用说明', '提示', 'readme', 'instruction')

    try:
        for ws in wb.worksheets:
            if any(kw in ws.title.lower() for kw in skip_sheet_keywords):
                continue

            # 扫描表头，定位各列
            header_map = {}
            header_row_idx = 0
            for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True)):
                row_strs = [str(c).strip() if c else '' for c in row]
                for ci, val in enumerate(row_strs):
                    if not val:
                        continue
                    val_lower = val.lower()
                    if val in ('项目名称', '项目', '名称', 'name', 'item'):
                        header_map['name'] = ci
                    elif val in ('病区', '病区/分类', '分类', 'area'):
                        header_map['area'] = ci
                    elif val in ('人次', 'session_count', '人次/数量', '数量'):
                        header_map['sessions'] = ci
                    elif val in ('金额', 'amount', '费用'):
                        header_map['amount'] = ci
                    elif val in ('月份', '月', 'month'):
                        header_map['month'] = ci
                if header_map:
                    header_row_idx = ri
                    break

            has_explicit_header = bool(header_map)
            name_col = header_map.get('name', 0)
            area_col = header_map.get('area')
            month_col = header_map.get('month')
            sessions_col = header_map.get('sessions')
            amount_col = header_map.get('amount')

            # 如果没有明确表头，用启发式检测双列
            has_two_cols = False
            if not has_explicit_header:
                for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
                    if row[0] and str(row[0]).strip():
                        if len(row) >= 2 and row[1] and str(row[1]).strip():
                            try: float(str(row[1]).strip())
                            except ValueError: has_two_cols = True
                        break

            skip_headers = ('项目名称', '项目', '名称', '病区', 'name', '病区/分类', '分类', '人次', '金额', '月份', '月', '数量', '费用', 'amount', 'month', 'area')

            current_area = None  # 继承上一行的病区
            for row in ws.iter_rows(values_only=True):
                if not row:
                    continue

                # 解析各列值
                first_val = str(row[0]).strip() if row[0] else ''
                second_val = str(row[1]).strip() if len(row) > 1 and row[1] else ''

                # 跳过表头行
                if first_val in skip_headers or second_val in skip_headers:
                    continue

                # 获取项目名称（优先B列）
                name = ''
                if has_explicit_header and name_col < len(row) and row[name_col]:
                    name = str(row[name_col]).strip()
                elif has_two_cols and not has_explicit_header and second_val:
                    name = second_val
                elif has_explicit_header and second_val:
                    name = second_val
                else:
                    name = first_val

                # 过滤纯数字
                if not name:
                    continue
                try:
                    float(name)
                    continue
                except ValueError:
                    pass

                # 获取病区（继承模式：空A列用上一个病区）
                # 注意：outpatient/children 等无子分类的类型，不要从A列自动推断病区
                if has_explicit_header and area_col is not None and area_col < len(row) and row[area_col]:
                    current_area = str(row[area_col]).strip()
                elif has_two_cols and not has_explicit_header and first_val:
                    # 无表头双列模式：A列=病区（仅病房等需要病区的类型）
                    if item_type in ('ward',):
                        current_area = first_val
                elif has_explicit_header and first_val and item_type in ('ward',):
                    # 有表头且是病房类型：A列有值可能是病区行
                    if second_val and second_val not in skip_headers:
                        current_area = first_val
                elif first_val and not second_val and item_type in ('ward',):
                    # 病房类型的标题行
                    current_area = first_val
                    continue

                # 门诊/儿童医院等无子分类的类型，强制area为空
                if item_type in ('outpatient', 'children'):
                    area = ''
                else:
                    area = current_area or sub
                # children_doctor等不需要病区的类型，允许area为空
                if not area and item_type not in ('children_doctor', 'children', 'outpatient', 'ward'):
                    continue

                # 获取月份
                row_month = None
                if month_col is not None and month_col < len(row) and row[month_col]:
                    try:
                        row_month = int(float(str(row[month_col]).strip()))
                        if row_month < 1 or row_month > 12:
                            row_month = None
                    except:
                        # 尝试解析 "X月" 格式
                        mstr = str(row[month_col]).strip()
                        import re
                        m = re.match(r'(\d+)', mstr)
                        if m:
                            row_month = int(m.group(1))
                            if row_month < 1 or row_month > 12:
                                row_month = None

                # 获取人次和金额
                sessions = 0
                amount = 0
                if sessions_col is not None and sessions_col < len(row) and row[sessions_col] is not None and str(row[sessions_col]).strip() != '':
                    try: sessions = int(float(str(row[sessions_col]).strip()))
                    except: pass
                if amount_col is not None and amount_col < len(row) and row[amount_col] is not None and str(row[amount_col]).strip() != '':
                    try: amount = float(str(row[amount_col]).strip())
                    except: pass

                # 无明确表头时，尝试固定位置读取
                if not has_explicit_header and sessions_col is None and amount_col is None:
                    data_start = 1 if not has_two_cols else 2
                    if len(row) > data_start and row[data_start] is not None and str(row[data_start]).strip() != '':
                        try: sessions = int(float(str(row[data_start]).strip()))
                        except: pass
                    if len(row) > data_start + 1 and row[data_start + 1] is not None and str(row[data_start + 1]).strip() != '':
                        try: amount = float(str(row[data_start + 1]).strip())
                        except: pass

                try:
                    # 导入项目到项目库
                    db.session.execute(text("""
                        UPDATE dept_items SET sort_order = :sort_order
                        WHERE item_type = :type AND item_name = :name AND sub_category = :sub
                    """), {'type': item_type, 'name': name, 'sub': area, 'sort_order': sort_idx})
                    db.session.execute(text("""
                        INSERT OR IGNORE INTO dept_items (item_type, item_name, sub_category, sort_order)
                        VALUES (:type, :name, :sub, :sort_order)
                    """), {'type': item_type, 'name': name, 'sub': area, 'sort_order': sort_idx})
                    sort_idx += 1
                    count += 1

                    # 写入月度统计（有选中月份就写入，允许人次数为0）
                    if months_list:
                        for mon in months_list:
                            try: mon = int(mon)
                            except: continue
                            if mon < 1 or mon > 12: continue
                            if item_type == 'outpatient':
                                db.session.execute(text("""
                                    INSERT INTO outpatient_monthly (item_name, year, month, session_count, amount)
                                    VALUES (:name, :year, :month, :sessions, :amount)
                                    ON CONFLICT(item_name, year, month) DO UPDATE SET
                                        session_count = :sessions, amount = :amount, updated_at = CURRENT_TIMESTAMP
                                """), {'name': name, 'year': int(year), 'month': mon,
                                      'sessions': sessions, 'amount': amount})
                                monthly_count += 1
                            elif item_type == 'children':
                                db.session.execute(text("""
                                    INSERT INTO children_monthly (item_name, year, month, session_count, amount)
                                    VALUES (:name, :year, :month, :sessions, :amount)
                                    ON CONFLICT(item_name, year, month) DO UPDATE SET
                                        session_count = COALESCE(:sessions, session_count), amount = COALESCE(:amount, amount), updated_at = CURRENT_TIMESTAMP
                                """), {'name': name, 'year': int(year), 'month': mon,
                                      'sessions': sessions, 'amount': amount})
                                monthly_count += 1
                            elif item_type == 'children_doctor':
                                db.session.execute(text("""
                                    INSERT INTO children_doctor_monthly (doctor_name, year, month, session_count)
                                    VALUES (:name, :year, :month, :sessions)
                                    ON CONFLICT(doctor_name, year, month) DO UPDATE SET
                                        session_count = :sessions, updated_at = CURRENT_TIMESTAMP
                                """), {'name': name, 'year': int(year), 'month': mon,
                                      'sessions': sessions})
                                monthly_count += 1
                            elif item_type == 'ward' and area:
                                db.session.execute(text("""
                                    INSERT INTO ward_monthly (ward_area, item_name, year, month, session_count, amount)
                                    VALUES (:area, :name, :year, :month, :sessions, :amount)
                                    ON CONFLICT(ward_area, item_name, year, month) DO UPDATE SET
                                        session_count = :sessions, amount = :amount, updated_at = CURRENT_TIMESTAMP
                                """), {'area': area, 'name': name, 'year': int(year), 'month': mon,
                                      'sessions': sessions, 'amount': amount})
                                monthly_count += 1
                except Exception as e:
                    errors.append(f'{name}: {str(e)}')

        db.session.commit()
        return jsonify({'success': True, 'data': {'count': count, 'monthly_count': monthly_count, 'errors': errors}})
    except Exception as e:
        import traceback as _tb
        db.session.rollback()
        _detail = _tb.format_exc()
        print('[IMPORT_ERROR] ' + str(e))
        print(_detail)
        return jsonify({'success': False, 'error': str(e)[:500], 'trace': _detail[-2000:]}), 500


@dept_stats_bp.route('/items/template', methods=['GET'])
def download_items_template():
    """下载项目导入模板"""
    wb = openpyxl.Workbook()
    item_type = request.args.get('type', 'outpatient')
    
    # 病房模板：按病区分组
    if item_type == 'ward':
        ws = wb.active
        ws.title = '病房项目库'
        
        # 表头
        headers = ['病区', '项目名称', '人次', '金额']
        header_fill = openpyxl.styles.PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        
        # 按病区分组
        ward_items = {
            '产科': ['低频', '气压', '产褥操', '微波', '生物反馈', '针灸', '推拿', '中药熏蒸', '红外线', '中药热敷'],
            '妇科': ['低频', '气压', '微波', '生物反馈', '针灸', '推拿', '红外线', '中药热敷', '盆底筛查', '盆底康复'],
            '产康中心': ['盆底磁', '盆底筛查', '腹直肌分离', '盆底康复', 'WA运动', '骨盆修复', '阴道哑铃训练', '电刺激', '生物反馈'],
            '外科': ['低频', '气压', '微波', '红外线', '中药热敷', '中药熏蒸'],
            '新生儿科': ['新生儿抚触', '新生儿游泳', '听力学筛查', '视力筛查', 'NBNA评分'],
        }
        row = 2
        for area, items in ward_items.items():
            for i, item in enumerate(items):
                ws.cell(row=row, column=1, value=area if i == 0 else '')
                ws.cell(row=row, column=2, value=item)
                row += 1
        
        # 合并病区名称单元格（先写值再合并）
        row = 2
        for area, items in ward_items.items():
            if len(items) > 1:
                ws.merge_cells(start_row=row, start_column=1, end_row=row + len(items) - 1, end_column=1)
                ws.cell(row=row, column=1).alignment = openpyxl.styles.Alignment(vertical='center', horizontal='center')
            row += len(items)
        
        # 边框样式
        thin_border = openpyxl.styles.Side(style='thin')
        border = openpyxl.styles.Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
        for r in range(1, row):
            for c in range(1, 5):
                cell = ws.cell(row=r, column=c)
                cell.border = border
        # 表头背景色
        header_fill = openpyxl.styles.PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        for c in range(1, 5):
            ws.cell(row=1, column=c).fill = header_fill
            ws.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        
        # 使用说明sheet
        ws2 = wb.create_sheet('填写说明')
        ws2.column_dimensions['A'].width = 60
        instructions = [
            '【填写说明】',
            '',
            '用法一：仅导入项目名称',
            '  A列填写病区，B列填写项目名称，C列留空，在「管理项目」或录入框中导入',
            '',
            '用法二：导入项目+月度数据',
            '  A列填写病区，B列填写项目名称，C列填写人次，D列填写金额',
            '  在录入框中选好年份和月份，然后点「导入月度数据」',
            '',
            '提示：系统会自动去重。月度数据如已存在会自动覆盖。',
        ]
        for i, line in enumerate(instructions, start=1):
            ws2.cell(row=i, column=1, value=line)
            if i == 1:
                ws2.cell(row=i, column=1).font = openpyxl.styles.Font(bold=True, size=14)
        
        fname = '病房项目导入模板.xlsx'
    
    # 门诊模板
    elif item_type == 'outpatient':
        ws = wb.active
        ws.title = '门诊项目库'
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        
        headers = ['项目名称', '人次', '金额']
        header_fill = openpyxl.styles.PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
        
        items = ['盆底筛查', '盆底磁（盆腔炎）', '腹直肌分离', '盆底康复', 'WA运动', '骨盆修复',
                 '阴道哑铃训练', '电刺激', '生物反馈', '产后塑形', '腹直肌评估',
                 '骨盆评估', '乳腺疏通', '催乳', '红外线', '微波', '低频', '气压',
                 '针灸', '推拿', '中药熏蒸', '中药热敷', '拔罐', '刮痧', '艾灸']
        for i, name in enumerate(items, start=2):
            ws.cell(row=i, column=1, value=name)
        
        thin_border = openpyxl.styles.Side(style='thin')
        border = openpyxl.styles.Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
        for r in range(1, len(items) + 2):
            for c in range(1, 4):
                ws.cell(row=r, column=c).border = border
        
        ws2 = wb.create_sheet('填写说明')
        ws2.column_dimensions['A'].width = 60
        for i, line in enumerate([
            '【填写说明】', '',
            '用法一：仅导入项目名称',
            '  A列填写项目名称，B列C列留空，在「管理项目」或录入框中导入即可',
            '',
            '用法二：导入项目+月度数据',
            '  A列填写项目名称，B列填写人次，C列填写金额',
            '  在录入框中选好年份和月份，然后点「导入月度数据」',
            '', '提示：系统会自动去重。月度数据如已存在会自动覆盖。'
        ], start=1):
            ws2.cell(row=i, column=1, value=line)
            if i == 1: ws2.cell(row=i, column=1).font = openpyxl.styles.Font(bold=True, size=14)
        
        fname = '门诊项目导入模板.xlsx'
    
    # 儿童医院模板 - 两个sheet：人次 & 金额
    elif item_type == 'children':
        items = ['PT', 'OT', 'ST', '感统', '小儿推拿', '儿童推拿', '言语/构音训练',
                 '经颅磁', '低磁', '运动发育迟缓训练', '肌张力异常训练',
                 '言语训练', '认知训练', '物理因子', '腹部/肺部物理治疗',
                 '幼儿斜颈', '高危儿早期干预', '步态训练', '作业治疗',
                 '感觉统合训练', '听力学评估']
        header_fill = openpyxl.styles.PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        thin_border = openpyxl.styles.Side(style='thin')
        border = openpyxl.styles.Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
        
        # Sheet 1: 人次
        ws1 = wb.active
        ws1.title = '人次'
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 10
        for ci, h in enumerate(['项目名称', '人次'], start=1):
            cell = ws1.cell(row=1, column=ci, value=h)
            cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
        for i, name in enumerate(items, start=2):
            ws1.cell(row=i, column=1, value=name)
        for r in range(1, len(items) + 2):
            for c in range(1, 3):
                ws1.cell(row=r, column=c).border = border
        
        # Sheet 2: 金额
        ws2 = wb.create_sheet('金额')
        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 12
        for ci, h in enumerate(['项目名称', '金额'], start=1):
            cell = ws2.cell(row=1, column=ci, value=h)
            cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
        for i, name in enumerate(items, start=2):
            ws2.cell(row=i, column=1, value=name)
        for r in range(1, len(items) + 2):
            for c in range(1, 3):
                ws2.cell(row=r, column=c).border = border
        
        # 填写说明
        ws3 = wb.create_sheet('填写说明')
        ws3.column_dimensions['A'].width = 60
        for i, line in enumerate([
            '【填写说明】', '',
            '「人次」sheet：A列填项目名称，B列填人次',
            '「金额」sheet：A列填项目名称，B列填金额',
            '',
            '用法一：仅导入项目名称',
            '  A列填项目名称，B列留空，在录入框中导入即可',
            '',
            '用法二：导入项目+月度数据',
            '  在录入框中选好年份和月份，然后点「导入月度数据」',
            '', '提示：系统会自动去重。月度数据如已存在会自动覆盖。'
        ], start=1):
            ws3.cell(row=i, column=1, value=line)
            if i == 1: ws3.cell(row=i, column=1).font = openpyxl.styles.Font(bold=True, size=14)
        
        fname = '儿童医院项目导入模板.xlsx'
    
    # 儿童医院医生开单模板
    elif item_type == 'children_doctor':
        ws = wb.active
        ws.title = '医生开单'
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 10
        header_fill = openpyxl.styles.PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        for ci, h in enumerate(['医生姓名', '人次'], start=1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
        thin_border = openpyxl.styles.Side(style='thin')
        border = openpyxl.styles.Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
        for r in range(1, 6):
            for c in range(1, 3):
                ws.cell(row=r, column=c).border = border
        ws2 = wb.create_sheet('填写说明')
        ws2.column_dimensions['A'].width = 60
        for i, line in enumerate([
            '【填写说明】', '',
            'A列填写医生姓名，B列填写人次',
            '',
            '用法一：仅导入医生姓名',
            '  A列填医生姓名，B列留空，在录入框中导入即可',
            '',
            '用法二：导入医生+月度数据',
            '  在录入框中选好年份和月份，然后点「导入Excel」',
            '', '提示：系统会自动去重。月度数据如已存在会自动覆盖。'
        ], start=1):
            ws2.cell(row=i, column=1, value=line)
            if i == 1: ws2.cell(row=i, column=1).font = openpyxl.styles.Font(bold=True, size=14)
        fname = '儿童医院医生开单导入模板.xlsx'
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=fname,
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# 转介医生/指标 Excel导入
@referral_bp.route('/doctors/import-excel', methods=['POST'])
def import_doctors_excel():
    """从Excel批量导入医生（支持单列或科室+医生双列）"""
    ensure_tables()
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '请上传文件'}), 400
    
    dept = request.form.get('department', '').strip()
    file = request.files['file']
    
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        count = 0
        skip = ('医生姓名', '医生', '姓名', 'name', '科室')
        
        for ws in wb.worksheets:
            has_two_cols = False
            for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
                if row[0] and str(row[0]).strip() not in skip and len(row) >= 2 and row[1] and str(row[1]).strip():
                    has_two_cols = True
                    break
            
            for row in ws.iter_rows(values_only=True):
                if not row or not row[0]: continue
                col_a = str(row[0]).strip() if row[0] else ''
                col_b = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                if col_a in skip or col_b in skip: continue
                
                if has_two_cols and col_b:
                    d, name = col_a, col_b
                else:
                    d, name = dept, col_a
                
                if not name: continue
                db.session.execute(text("""
                    INSERT OR IGNORE INTO referral_doctors (department, doctor_name) VALUES (:dept, :name)
                """), {'dept': d or dept, 'name': name})
                count += 1
        db.session.commit()
        return jsonify({'success': True, 'data': {'count': count}})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


def import_metrics_excel():
    """从Excel批量导入指标（支持单列或科室+指标双列）"""
    ensure_tables()
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '请上传文件'}), 400
    
    dept = request.form.get('department', '').strip()
    file = request.files['file']
    
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        count = 0
        skip = ('指标名称', '指标', '名称', '科室')
        
        for ws in wb.worksheets:
            has_two_cols = False
            for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
                if row[0] and str(row[0]).strip() not in skip and len(row) >= 2 and row[1] and str(row[1]).strip():
                    has_two_cols = True
                    break
            
            for row in ws.iter_rows(values_only=True):
                if not row or not row[0]: continue
                col_a = str(row[0]).strip() if row[0] else ''
                col_b = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                if col_a in skip or col_b in skip: continue
                
                if has_two_cols and col_b:
                    d, name = col_a, col_b
                else:
                    d, name = dept, col_a
                
                if not name: continue
                db.session.execute(text("""
                    INSERT OR IGNORE INTO referral_metrics (department, metric_name) VALUES (:dept, :name)
                """), {'dept': d or dept, 'name': name})
                count += 1
        db.session.commit()
        return jsonify({'success': True, 'data': {'count': count}})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500




# ============================================================================
# 二、门诊数据
# ============================================================================

@dept_stats_bp.route('/outpatient', methods=['GET'])
def get_outpatient():
    """获取门诊月度数据（支持按月筛选）"""
    ensure_tables()
    year = request.args.get('year', date.today().year, type=int)
    month = request.args.get('month', type=int)

    sql = "SELECT item_name, month, session_count, amount FROM outpatient_monthly WHERE year = :year"
    params = {'year': year}
    if month:
        sql += " AND month = :month"
        params['month'] = month
    sql += " ORDER BY month, item_name"

    rows = db.session.execute(text(sql), params).fetchall()

    return jsonify({'success': True, 'data': [
        {'item': r[0], 'month': r[1], 'sessions': r[2] or 0, 'amount': r[3] or 0}
        for r in rows
    ]})


@dept_stats_bp.route('/outpatient/monthly-data', methods=['GET'])
def get_outpatient_monthly_data():
    """获取门诊某年某月的已有数据（用于录入框回填）"""
    ensure_tables()
    year = request.args.get('year', '')
    month = request.args.get('month', '')
    if not year or not month:
        return jsonify({'success': True, 'data': {}})
    rows = db.session.execute(text("""
        SELECT item_name, session_count, amount FROM outpatient_monthly
        WHERE year = :year AND month = :month
    """), {'year': int(year), 'month': int(month)}).fetchall()
    data = {r[0]: {'session_count': r[1] or 0, 'amount': r[2] or 0} for r in rows}
    return jsonify({'success': True, 'data': data})


@dept_stats_bp.route('/outpatient', methods=['POST'])
def save_outpatient():
    """保存门诊月度数据（批量，upsert，保存前自动快照）"""
    ensure_tables()
    data = request.get_json()
    year = data.get('year')
    month = data.get('month')
    records = data.get('records', [])
    trigger_type = data.get('trigger_type', 'auto_save')

    if not year or not month or not records:
        return jsonify({'success': False, 'error': '缺少必要参数'}), 400

    # 自动快照
    try:
        _create_snapshot('outpatient', f'{year}-{month}', 'outpatient_monthly',
            ['item_name', 'year', 'month'],
            "SELECT * FROM outpatient_monthly WHERE year=:y AND month=:m",
            {'y': int(year), 'm': int(month)}, trigger_type=trigger_type)
    except Exception:
        pass

    count = 0
    for r in records:
        name = r.get('item_name', '').strip()
        sessions = r.get('session_count', 0)
        amount = r.get('amount', 0)
        if not name:
            continue
        
        db.session.execute(text("""
            INSERT INTO outpatient_monthly (item_name, year, month, session_count, amount)
            VALUES (:name, :year, :month, :sessions, :amount)
            ON CONFLICT(item_name, year, month) DO UPDATE SET
                session_count = :sessions, amount = :amount, updated_at = CURRENT_TIMESTAMP
        """), {'name': name, 'year': int(year), 'month': int(month),
              'sessions': int(sessions), 'amount': float(amount)})
        count += 1
    
    db.session.commit()
    return jsonify({'success': True, 'data': {'count': count}})


@dept_stats_bp.route('/outpatient/<int:record_id>', methods=['DELETE'])
def delete_outpatient(record_id):
    ensure_tables()
    db.session.execute(text("DELETE FROM outpatient_monthly WHERE id = :id"), {'id': record_id})
    db.session.commit()
    return jsonify({'success': True})


@dept_stats_bp.route('/outpatient/clear', methods=['POST'])
def clear_outpatient():
    ensure_tables()
    data = request.get_json() or {}
    year = data.get('year')
    month = data.get('month')
    if year and month:
        db.session.execute(text("DELETE FROM outpatient_monthly WHERE year = :y AND month = :m"),
                         {'y': int(year), 'm': int(month)})
    else:
        db.session.execute(text("DELETE FROM outpatient_monthly"))
    db.session.commit()
    return jsonify({'success': True})


# ============================================================================
# 三、病房数据
# ============================================================================

@dept_stats_bp.route('/ward', methods=['GET'])
def get_ward():
    ensure_tables()
    year = request.args.get('year', date.today().year, type=int)
    
    rows = db.session.execute(text("""
        SELECT id, ward_area, item_name, month, session_count, amount
        FROM ward_monthly WHERE year = :year
        ORDER BY ward_area, month, item_name
    """), {'year': year}).fetchall()

    return jsonify({'success': True, 'data': [
        {'id': r[0], 'area': r[1], 'item': r[2], 'month': r[3], 'sessions': r[4] or 0, 'amount': r[5] or 0}
        for r in rows
    ]})


@dept_stats_bp.route('/ward', methods=['POST'])
def save_ward():
    """保存病房月度数据（批量，upsert，保存前自动快照）"""
    ensure_tables()
    data = request.get_json()
    year = data.get('year')
    month = data.get('month')
    records = data.get('records', [])
    trigger_type = data.get('trigger_type', 'auto_save')

    if not year or not month or not records:
        return jsonify({'success': False, 'error': '缺少必要参数'}), 400

    # 自动快照
    try:
        _create_snapshot('ward_monthly', f'{year}-{month}', 'ward_monthly',
            ['ward_area', 'item_name', 'year', 'month'],
            "SELECT * FROM ward_monthly WHERE year=:y AND month=:m",
            {'y': int(year), 'm': int(month)}, trigger_type=trigger_type)
    except Exception:
        pass

    count = 0
    for r in records:
        area = r.get('ward_area', '').strip()
        name = r.get('item_name', '').strip()
        sessions = r.get('session_count', 0)
        amount = r.get('amount', 0)
        if not area or not name:
            continue
        
        db.session.execute(text("""
            INSERT INTO ward_monthly (ward_area, item_name, year, month, session_count, amount)
            VALUES (:area, :name, :year, :month, :sessions, :amount)
            ON CONFLICT(ward_area, item_name, year, month) DO UPDATE SET
                session_count = :sessions, amount = :amount, updated_at = CURRENT_TIMESTAMP
        """), {'area': area, 'name': name, 'year': int(year), 'month': int(month),
              'sessions': int(sessions), 'amount': float(amount)})
        count += 1
    
    db.session.commit()
    return jsonify({'success': True, 'data': {'count': count}})


@dept_stats_bp.route('/ward/<int:record_id>', methods=['DELETE'])
def delete_ward(record_id):
    ensure_tables()
    db.session.execute(text("DELETE FROM ward_monthly WHERE id = :id"), {'id': record_id})
    db.session.commit()
    return jsonify({'success': True})

@dept_stats_bp.route('/outpatient/by-item-month', methods=['DELETE'])
def delete_outpatient_by_item_month():
    """按项目名+月份删除门诊数据"""
    ensure_tables()
    data = request.get_json() or {}
    item = data.get('item_name', '')
    month = int(data.get('month', 0))
    if not item or not month:
        return jsonify({'success': False, 'error': '参数不完整'}), 400
    db.session.execute(text("DELETE FROM outpatient_monthly WHERE item_name = :item AND month = :month"),
                 {'item': item, 'month': month})
    db.session.commit()
    from app.api.operation_log_bp import log_operation
    log_operation('delete', f'删除门诊数据: {item} {month}月')
    return jsonify({'success': True})


@dept_stats_bp.route('/outpatient/batch-delete', methods=['POST'])
def batch_delete_outpatient():
    ensure_tables()
    data = request.get_json() or {}
    records = data.get('records', [])
    if not records:
        return jsonify({'success': False, 'error': '请提供要删除的记录'}), 400
    deleted = 0
    for rec in records:
        db.session.execute(text("DELETE FROM outpatient_monthly WHERE item_name = :item AND month = :month"),
                     {'item': rec.get('item_name',''), 'month': int(rec.get('month',0))})
        deleted += 1
    db.session.commit()
    from app.api.operation_log_bp import log_operation
    log_operation('delete', f'批量删除门诊数据 {deleted}条')
    return jsonify({'success': True, 'deleted': deleted})


@dept_stats_bp.route('/ward/batch-delete', methods=['POST'])
def batch_delete_ward():
    ensure_tables()
    data = request.get_json() or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'error': '请提供要删除的ID'}), 400
    deleted = 0
    for rid in ids:
        db.session.execute(text("DELETE FROM ward_monthly WHERE id = :id"), {'id': int(rid)})
        deleted += 1
    db.session.commit()
    from app.api.operation_log_bp import log_operation
    log_operation('delete', f'批量删除病房数据 {deleted}条')
    return jsonify({'success': True, 'deleted': deleted})


@dept_stats_bp.route('/referral/doctors/batch-delete', methods=['POST'])
def batch_delete_referral_doctors():
    ensure_tables()
    data = request.get_json() or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'error': '请提供要删除的ID'}), 400
    deleted = 0
    for did in ids:
        row = db.session.execute(text("SELECT department, doctor_name FROM referral_doctors WHERE id = :id"), {'id': int(did)}).fetchone()
        if row:
            dept, name = row[0], row[1]
            db.session.execute(text("DELETE FROM referral_monthly WHERE department = :dept AND doctor_name = :doc"),
                         {'dept': dept, 'doc': name})
            db.session.execute(text("DELETE FROM referral_doctors WHERE id = :id"), {'id': int(did)})
            deleted += 1
    db.session.commit()
    from app.api.operation_log_bp import log_operation
    log_operation('delete', f'批量删除转介数据 {deleted}条')
    return jsonify({'success': True, 'deleted': deleted})




@dept_stats_bp.route('/ward/clear', methods=['POST'])
def clear_ward():
    ensure_tables()
    data = request.get_json() or {}
    year = data.get('year')
    month = data.get('month')
    if year and month:
        db.session.execute(text("DELETE FROM ward_monthly WHERE year = :y AND month = :m"),
                         {'y': int(year), 'm': int(month)})
    else:
        db.session.execute(text("DELETE FROM ward_monthly"))
    db.session.commit()
    return jsonify({'success': True})


@dept_stats_bp.route('/ward/areas', methods=['GET'])
def get_ward_areas():
    """获取病区列表"""
    ensure_tables()
    rows = db.session.execute(text("""
        SELECT DISTINCT sub_category FROM dept_items
        WHERE item_type = 'ward' AND sub_category != ''
        ORDER BY sub_category
    """)).fetchall()
    return jsonify({'success': True, 'data': [r[0] for r in rows]})


# ============================================================================
# 三点五、病房每日统计（透视表录入）
# ============================================================================

_DEFAULT_WARD_PRICES = [
    ('低频8部位', 136), ('低频6部位', 102), ('气压', 34),
    ('康复评定', 34), ('通乳', 101.2), ('脏腑推拿', 42.3),
    ('短波', 21.6), ('耻骨', 207.8), ('会阴', 114),
    ('热奄包', 20.2), ('中药贴敷', 40.4), ('耳穴疗法', 40.4),
    ('紫外线', 10),
]


def _seed_ward_daily_defaults():
    """种子默认项目和单价（幂等，INSERT OR IGNORE）"""
    area = ''
    for name, price in _DEFAULT_WARD_PRICES:
        db.session.execute(text("""
            INSERT OR IGNORE INTO ward_item_prices (ward_area, item_name, unit_price)
            VALUES (:area, :name, :price)
        """), {'area': area, 'name': name, 'price': price})
    db.session.commit()


@dept_stats_bp.route('/ward/daily/prices', methods=['GET'])
def get_ward_daily_prices():
    """获取病房项目单价列表"""
    ensure_tables()
    area = request.args.get('area', '')
    query = "SELECT item_name, unit_price, sort_order FROM ward_item_prices WHERE is_active = 1"
    params = {}
    if area:
        query += " AND ward_area = :area"
        params['area'] = area
    query += " ORDER BY sort_order, item_name"
    rows = db.session.execute(text(query), params).fetchall()
    # 如果没有数据，先种子
    if not rows:
        _seed_ward_daily_defaults()
        rows = db.session.execute(text(query), params).fetchall()
    return jsonify({'success': True, 'data': [
        {'item_name': r[0], 'unit_price': float(r[1]), 'sort_order': int(r[2])} for r in rows
    ]})


@dept_stats_bp.route('/ward/daily/prices', methods=['POST'])
def save_ward_daily_prices():
    """批量保存单价（含增删同步）"""
    ensure_tables()
    data = request.get_json()
    area = data.get('area', '')
    prices = data.get('prices', [])
    # Collect names being saved
    kept_names = set()
    for p in prices:
        name = (p.get('item_name') or '').strip()
        price = float(p.get('unit_price') or 0)
        sort_order = int(p.get('sort_order', 0))
        if not name:
            continue
        kept_names.add(name)
        db.session.execute(text("""
            INSERT INTO ward_item_prices (ward_area, item_name, unit_price, sort_order)
            VALUES (:area, :name, :price, :sort_order)
            ON CONFLICT(ward_area, item_name) DO UPDATE SET
                unit_price = :price, sort_order = :sort_order, updated_at = CURRENT_TIMESTAMP
        """), {'area': area, 'name': name, 'price': price, 'sort_order': sort_order})
    # Delete items that were removed from the list
    if area and kept_names:
        placeholders = ','.join([f':n{i}' for i in range(len(kept_names))])
        params = {**{f'n{i}': n for i, n in enumerate(kept_names)}, 'area': area}
        db.session.execute(text(f"""
            DELETE FROM ward_item_prices WHERE ward_area = :area AND item_name NOT IN ({placeholders})
        """), params)
    elif area and not kept_names:
        # If all items removed, clear all prices for this area
        db.session.execute(text("DELETE FROM ward_item_prices WHERE ward_area = :area"), {'area': area})
    db.session.commit()
    return jsonify({'success': True})


@dept_stats_bp.route('/ward/daily/prices/seed', methods=['POST'])
def seed_ward_daily_prices():
    """手动触发种子默认数据"""
    ensure_tables()
    _seed_ward_daily_defaults()
    return jsonify({'success': True})


# ============================================================================
# 通用项目管理（门诊/儿童/转介 的项目编排 + 单价）
# ============================================================================

@dept_stats_bp.route('/items/manage', methods=['GET'])
def get_module_items():
    """获取某模块的项目列表（含排序和分类），为空时自动从数据表初始化"""
    ensure_tables()
    item_type = request.args.get('type', '').strip()
    if not item_type:
        return jsonify({'success': False, 'error': '缺少type参数'}), 400
    rows = db.session.execute(text("""
        SELECT item_name, sub_category, sort_order FROM dept_items
        WHERE item_type = :itype AND is_active = 1
        ORDER BY sort_order, item_name
    """), {'itype': item_type}).fetchall()

    # 为空时自动从对应数据表初始化项目列表
    if not rows:
        _auto_seed_module_items(item_type)
        rows = db.session.execute(text("""
            SELECT item_name, sub_category, sort_order FROM dept_items
            WHERE item_type = :itype AND is_active = 1
            ORDER BY sort_order, item_name
        """), {'itype': item_type}).fetchall()

    # 去重：同名项目只保留 sort_order 最小的一条
    seen = {}
    unique_rows = []
    for r in rows:
        name = r[0]
        if name not in seen:
            seen[name] = True
            unique_rows.append(r)

    return jsonify({'success': True, 'data': [
        {'item_name': r[0], 'sub_category': r[1] or '', 'sort_order': int(r[2] or 0)}
        for r in unique_rows
    ]})


def _auto_seed_module_items(item_type):
    """从对应数据表自动提取项目名，初始化到 dept_items（幂等，仅提取当前有效数据）"""
    # 无效名称黑名单（汇总行、测试数据、空值变体）
    _invalid_names = {'合计', '总计', '小计', '日期', '(未分类)', '-', '',
                      '测试', '新项目a', '新项目b', '新项目c',
                      '新项目A', '新项目B', '新项目C', '新项目D',
                      '测试科', '测试科室', 'test'}
    try:
        if item_type == 'referral':
            # 转介：从 referral_doctors 表取医生（按科室分组）
            rows = db.session.execute(text("""
                SELECT doctor_name, department FROM referral_doctors
                WHERE is_active = 1 AND doctor_name IS NOT NULL AND doctor_name != ''
                ORDER BY department, doctor_name
            """)).fetchall()
            if not rows:
                return
            idx = 0
            for r in rows:
                name = (r[0] or '').strip()  # doctor_name (第1列)
                dept = (r[1] or '').strip()   # department → sub_category (第2列)
                if not name or name in _invalid_names:
                    continue
                # 额外过滤：含"测试"、纯数字、过短
                if '测试' in name or name.isdigit() or len(name) < 2:
                    continue
                db.session.execute(text("""
                    INSERT OR IGNORE INTO dept_items (item_type, item_name, sub_category, unit_price, sort_order)
                    VALUES (:itype, :name, :subcat, 0, :sort)
                """), {'itype': item_type, 'name': name, 'subcat': dept, 'sort': idx})
                idx += 1
            db.session.commit()
            return

        # 其他模块：从月度数据表取项目名
        table_map = {
            'outpatient': ('outpatient_monthly', 'item_name'),
            'children': ('children_monthly', 'item_name'),
            'children_doctor': ('children_doctor_monthly', "COALESCE(metric_name,doctor_name||'')"),
        }
        if item_type not in table_map:
            return
        table_name, name_col = table_map[item_type]
        # 只取当前年份的有效数据，避免已删除的陈旧记录被重新引入
        cur_year = datetime.now().year
        rows = db.session.execute(text(f"""
            SELECT DISTINCT {name_col} AS item_name FROM {table_name}
            WHERE {name_col} IS NOT NULL AND {name_col} != ''
              AND year = :cur_year
            ORDER BY {name_col}
        """), {'cur_year': cur_year}).fetchall()
        if not rows:
            return
        idx = 0
        for r in rows:
            name = (r[0] or '').strip()
            if not name or name in _invalid_names:
                continue
            if name.isdigit() or len(name) < 2:
                continue
            db.session.execute(text("""
                INSERT OR IGNORE INTO dept_items (item_type, item_name, sub_category, unit_price, sort_order)
                VALUES (:itype, :name, '', 0, :sort)
            """), {'itype': item_type, 'name': name, 'sort': idx})
            idx += 1
        db.session.commit()
    except Exception:
        db.session.rollback()


@dept_stats_bp.route('/items/manage', methods=['POST'])
def save_module_items():
    """批量保存模块项目（全量替换：先清空再插入，避免重复）"""
    ensure_tables()
    data = request.get_json() or {}
    item_type = data.get('type', '').strip()
    items = data.get('items', [])
    if not item_type:
        return jsonify({'success': False, 'error': '缺少type参数'}), 400
    # 先清空该类型的所有旧数据（避免重复）
    db.session.execute(text("DELETE FROM dept_items WHERE item_type = :itype"), {'itype': item_type})
    # 全量插入新数据
    for idx, it in enumerate(items):
        name = (it.get('name') or it.get('item_name') or '').strip()
        subcat = (it.get('sub_category') or '').strip()
        if not name:
            continue
        db.session.execute(text("""
            INSERT INTO dept_items (item_type, item_name, sub_category, unit_price, sort_order, is_active)
            VALUES (:itype, :name, :subcat, 0, :sort, 1)
        """), {'itype': item_type, 'name': name, 'subcat': subcat, 'sort': idx})
    db.session.commit()
    return jsonify({'success': True, 'data': {'count': len(items)}})


@dept_stats_bp.route('/ward/daily', methods=['GET'])
def get_ward_daily():
    """查询每日数据"""
    ensure_tables()
    year = request.args.get('year', '')
    month = request.args.get('month', '')
    area = request.args.get('area', '')

    conditions = []
    params = {}
    if year:
        conditions.append("strftime('%Y', record_date) = :year")
        params['year'] = str(year)
    if month:
        conditions.append("CAST(strftime('%m', record_date) AS INTEGER) = :month")
        params['month'] = int(month)
    if area:
        conditions.append("ward_area = :area")
        params['area'] = area

    where = " WHERE " + " AND ".join(conditions) if conditions else ""
    rows = db.session.execute(text(f"""
        SELECT record_date, ward_area, item_name, session_count, amount
        FROM ward_daily{where}
        ORDER BY record_date, item_name
    """), params).fetchall()

    return jsonify({'success': True, 'data': [
        {
            'date': str(r[0]),
            'area': r[1],
            'item': r[2],
            'sessions': r[3] or 0,
            'amount': float(r[4]) if r[4] is not None else 0.0
        } for r in rows
    ]})


def _create_ward_daily_snapshot(area, trigger_type='auto_save'):
    """保存前快照当前area的所有数据，用于回滚"""
    rows = db.session.execute(text("""
        SELECT record_date, item_name, session_count, amount
        FROM ward_daily WHERE ward_area = :area ORDER BY record_date, item_name
    """), {'area': area}).fetchall()
    if not rows:
        return
    snapshot_data = [
        {'record_date': str(r[0]), 'item_name': r[1], 'session_count': r[2] or 0,
         'amount': float(r[3]) if r[3] is not None else 0.0}
        for r in rows
    ]
    _now_local = "datetime('now','localtime')"
    # import/sync 类型：始终创建新快照（重要回滚点，不节流）
    if trigger_type in ('import', 'sync'):
        db.session.execute(text("""
            INSERT INTO ward_daily_snapshots (ward_area, trigger_type, data_json, record_count, snapshot_time)
            VALUES (:area, :type, :dj, :rc, """ + _now_local + """)
        """), {'area': area, 'type': trigger_type,
              'dj': json.dumps(snapshot_data, ensure_ascii=False),
              'rc': len(snapshot_data)})
        _prune_snapshots('ward_daily_snapshots', f'ward_area = "{area}"')
        return
    # auto_save：30分钟内只创建1次（不覆盖已有快照，保留历史版本）
    if trigger_type == 'auto_save':
        recent = db.session.execute(text("""
            SELECT id FROM ward_daily_snapshots
            WHERE ward_area = :area AND trigger_type = 'auto_save'
              AND snapshot_time > """ + _now_local + """ || '-30 minutes'
            ORDER BY id DESC LIMIT 1
        """), {'area': area}).fetchone()
        if recent:
            return  # 30分钟内有快照了，跳过（不覆盖，保留之前的版本）
    db.session.execute(text("""
        INSERT INTO ward_daily_snapshots (ward_area, trigger_type, data_json, record_count, snapshot_time)
        VALUES (:area, :type, :dj, :rc, """ + _now_local + """)
    """), {'area': area, 'type': trigger_type,
          'dj': json.dumps(snapshot_data, ensure_ascii=False),
          'rc': len(snapshot_data)})
    _prune_snapshots('ward_daily_snapshots', f'ward_area = "{area}"')


# ============================================================================
# 通用数据快照工具函数（供各模块回滚使用）
# ============================================================================

def _create_snapshot(module, scope_key, table_name, key_columns, query_sql,
                     params, trigger_type='auto_save'):
    """通用快照：查询当前数据→序列化JSON→存入data_snapshots表（含节流）"""
    result = db.session.execute(text(query_sql), params)
    cols = list(result.keys()) if result else []
    rows = result.fetchall()
    if not rows:
        return
    snapshot_data = [dict(zip(cols, r)) for r in rows]
    _now_local = "datetime('now','localtime')"
    # import/sync 类型：始终创建新快照（重要回滚点，不节流）
    if trigger_type in ('import', 'sync'):
        db.session.execute(text("""
            INSERT INTO data_snapshots (module, scope_key, table_name, key_columns,
                                        trigger_type, data_json, record_count, snapshot_time)
            VALUES (:mod, :sk, :tbl, :kc, :type, :dj, :rc, """ + _now_local + """)
        """), {'mod': module, 'sk': scope_key, 'tbl': table_name,
              'kc': json.dumps(key_columns), 'type': trigger_type,
              'dj': json.dumps(snapshot_data, ensure_ascii=False),
              'rc': len(snapshot_data)})
        _prune_snapshots('data_snapshots', f'module = "{module}"')
        return
    # auto_save：30分钟内只创建1次（不覆盖已有快照，保留历史版本）
    if trigger_type == 'auto_save':
        recent = db.session.execute(text("""
            SELECT id FROM data_snapshots
            WHERE module=:mod AND scope_key=:sk AND trigger_type='auto_save'
              AND snapshot_time > """ + _now_local + """ || '-30 minutes'
            ORDER BY id DESC LIMIT 1
        """), {'mod': module, 'sk': scope_key}).fetchone()
        if recent:
            return  # 30分钟内有快照了，跳过
    db.session.execute(text("""
        INSERT INTO data_snapshots (module, scope_key, table_name, key_columns,
                                    trigger_type, data_json, record_count, snapshot_time)
        VALUES (:mod, :sk, :tbl, :kc, :type, :dj, :rc, """ + _now_local + """)
    """), {'mod': module, 'sk': scope_key, 'tbl': table_name,
          'kc': json.dumps(key_columns), 'type': trigger_type,
          'dj': json.dumps(snapshot_data, ensure_ascii=False),
          'rc': len(snapshot_data)})
    _prune_snapshots('data_snapshots', f'module = "{module}"')


def _prune_snapshots(table_name, where_clause):
    """清理快照表：每个分组最多保留50条（保留最近的，删除最旧的）"""
    try:
        # 查询总数
        total = db.session.execute(text(
            f"SELECT COUNT(*) FROM {table_name} WHERE {where_clause}"
        )).scalar() or 0
        if total <= 50:
            return
        # 删除超出部分（保留最新的50条）
        keep_ids = db.session.execute(text(
            f"SELECT id FROM {table_name} WHERE {where_clause} "
            f"ORDER BY id DESC LIMIT 50"
        )).fetchall()
        if not keep_ids:
            return
        keep_str = ','.join(str(r[0]) for r in keep_ids)
        deleted = db.session.execute(text(
            f"DELETE FROM {table_name} WHERE {where_clause} AND id NOT IN ({keep_str})"
        )).rowcount
    except Exception:
        pass  # 清理失败不应阻断主流程


def _do_rollback(snapshot_id):
    """通用回滚：从data_snapshots恢复数据到原表"""
    row = db.session.execute(text(
        "SELECT table_name, key_columns, data_json FROM data_snapshots WHERE id = :sid"
    ), {'sid': snapshot_id}).fetchone()
    if not row:
        return None, '快照不存在'
    table_name = row[0]
    key_cols = json.loads(row[1])
    records = json.loads(row[2])
    count = 0
    for rec in records:
        # 构建 SET 子句（排除key列）
        set_parts = [f"{k}=:{k}" for k in rec.keys() if k not in key_cols and k != 'id']
        # 构建 VALUES/ON CONFLICT
        col_names = list(rec.keys())
        placeholders = [f":{c}" for c in col_names]
        all_params = {c: rec[c] for c in col_names}
        # UPSERT
        sql = f"INSERT INTO {table_name} ({', '.join(col_names)}) VALUES ({', '.join(placeholders)})"
        if key_cols:
            on_conflict = f" ON CONFLICT({', '.join(key_cols)}) DO UPDATE SET {', '.join(set_parts)}"
            # 为 DO UPDATE 部分添加排除参数
            for kc in key_cols:
                if kc in rec:
                    pass  # key列用 excluded. 引用或直接值
            sql += on_conflict
            # 合并参数：SET部分需要引用excluded的值
            for sk in set_parts:
                col_name = sk.split('=')[0].strip()
                if col_name in rec:
                    all_params[f'excluded_{col_name}'] = rec[col_name]
        try:
            db.session.execute(text(sql), all_params)
            count += 1
        except Exception:
            pass
    db.session.commit()
    return count, None


@dept_stats_bp.route('/ward/daily', methods=['POST'])
def save_ward_daily():
    """批量保存每日记录（支持auto_calc自动计算金额，保存前自动快照）"""
    ensure_tables()
    data = request.get_json()
    records = data.get('records', [])
    auto_calc = data.get('auto_calc', True)
    trigger_type = data.get('trigger_type', 'auto_save')  # auto_save / sync / import

    # 可选：同时更新单价（按各记录的ward_area分别保存）
    if data.get('update_prices') and data.get('prices'):
        # 收集本次请求涉及的所有area
        areas_in_request = set()
        for r in records:
            a = str(r.get('ward_area') or '').strip()
            if a: areas_in_request.add(a)
        for p in data['prices']:
            name = (p.get('item_name') or '').strip()
            price = float(p.get('unit_price') or 0)
            if name:
                # 为每个涉及的area都写入价格
                for area in (areas_in_request or ['']):
                    db.session.execute(text("""
                        INSERT INTO ward_item_prices (ward_area, item_name, unit_price)
                        VALUES (:area, :name, :price)
                        ON CONFLICT(ward_area, item_name) DO UPDATE SET
                            unit_price = :price, updated_at = CURRENT_TIMESTAMP
                    """), {'area': area, 'name': name, 'price': price})

    # 自动快照：在覆盖数据前记录当前状态（用于回滚）
    if records:
        areas_in_request = set()
        for r in records:
            a = str(r.get('ward_area') or '').strip() or ''
            if a: areas_in_request.add(a)
        for area in areas_in_request:
            try: _create_ward_daily_snapshot(area, trigger_type)
            except Exception: pass  # 快照失败不应阻断保存

    count = 0
    # 预加载单价缓存
    price_cache = {}
    if auto_calc:
        pr = db.session.execute(text(
            "SELECT item_name, unit_price FROM ward_item_prices WHERE is_active=1"
        )).fetchall()
        price_cache = {r[0]: float(r[1]) for r in pr}

    for r in records:
        raw_date = str(r.get('record_date') or '').strip()
        # Convert Excel serial dates (e.g. 46054 → 2026-02-01)
        date_val = raw_date
        if raw_date and re.match(r'^\d{5,}$', raw_date):
            try:
                serial = int(raw_date)
                dt = datetime(1899, 12, 30) + timedelta(days=serial)
                date_val = dt.strftime('%Y-%m-%d')
            except (ValueError, OverflowError):
                pass
        area = str(r.get('ward_area') or '').strip() or ''
        name = str(r.get('item_name') or '').strip()
        sessions = int(r.get('session_count') or 0)

        if not date_val or not name:
            continue

        if auto_calc:
            unit_price = price_cache.get(name, 0)
            amt = round(sessions * unit_price, 2)
        else:
            amt = float(r.get('amount') or 0)

        db.session.execute(text("""
            INSERT INTO ward_daily (record_date, ward_area, item_name, session_count, amount)
            VALUES (:date, :area, :name, :sessions, :amt)
            ON CONFLICT(record_date, ward_area, item_name) DO UPDATE SET
                session_count = :sessions, amount = :amt, updated_at = CURRENT_TIMESTAMP
        """), {'date': date_val, 'area': area, 'name': name,
              'sessions': sessions, 'amt': amt})
        count += 1

    db.session.commit()
    return jsonify({'success': True, 'data': {'count': count}})


@dept_stats_bp.route('/ward/daily', methods=['DELETE'])
def delete_ward_daily():
    """删除某日数据"""
    ensure_tables()
    date_val = request.args.get('date', '')
    area = request.args.get('area', '')
    if not date_val:
        return jsonify({'success': False, 'error': '缺少日期参数'}), 400
    sql = "DELETE FROM ward_daily WHERE record_date = :date"
    params = {'date': date_val}
    if area:
        sql += " AND ward_area = :area"
        params['area'] = area
    db.session.execute(text(sql), params)
    db.session.commit()
    return jsonify({'success': True})


@dept_stats_bp.route('/ward/daily/clear', methods=['POST'])
def clear_ward_daily():
    """清除月/年数据"""
    ensure_tables()
    data = request.get_json() or {}
    year = data.get('year')
    month = data.get('month')

    if year and month:
        db.session.execute(text("""
            DELETE FROM ward_daily
            WHERE strftime('%Y', record_date) = :y
              AND CAST(strftime('%m', record_date) AS INTEGER) = :m
        """), {'y': str(year), 'm': int(month)})
    elif year:
        db.session.execute(text("""
            DELETE FROM ward_daily WHERE strftime('%Y', record_date) = :y
        """), {'y': str(year)})
    else:
        db.session.execute(text("DELETE FROM ward_daily"))
    db.session.commit()
    return jsonify({'success': True})


@dept_stats_bp.route('/ward/daily/rollup', methods=['POST'])
def rollup_ward_daily_to_monthly():
    """将每日数据汇总到月度表"""
    ensure_tables()
    data = request.get_json() or {}
    year = data.get('year')
    month = data.get('month')

    if not year or not month:
        return jsonify({'success': False, 'error': '缺少年月参数'}), 400

    db.session.execute(text("""
        INSERT INTO ward_monthly (ward_area, item_name, year, month, session_count, amount)
        SELECT ward_area, item_name, :year, :month,
               COALESCE(SUM(session_count), 0),
               COALESCE(SUM(amount), 0)
        FROM ward_daily
        WHERE strftime('%Y', record_date) = :y
          AND CAST(strftime('%m', record_date) AS INTEGER) = :m
        GROUP BY ward_area, item_name
        ON CONFLICT(ward_area, item_name, year, month) DO UPDATE SET
            session_count = excluded.session_count,
            amount = excluded.amount,
            updated_at = CURRENT_TIMESTAMP
    """), {'y': str(year), 'm': int(month), 'year': int(year), 'month': int(month)})
    db.session.commit()
    return jsonify({'success': True})


@dept_stats_bp.route('/ward/daily/monthly-summary', methods=['GET'])
def get_ward_daily_monthly_summary():
    """从ward_daily聚合月度汇总（按科室+项目分组）"""
    ensure_tables()
    year = request.args.get('year', date.today().year, type=int)
    month = request.args.get('month', type=int)
    area = request.args.get('area', '').strip()

    query = """
        SELECT ward_area, item_name,
               COALESCE(SUM(session_count), 0) as total_sessions,
               COALESCE(SUM(amount), 0) as total_amount,
               COUNT(DISTINCT record_date) as day_count
        FROM ward_daily
        WHERE strftime('%Y', record_date) = :y
    """
    params = {'y': str(year)}
    if month:
        query += " AND CAST(strftime('%m', record_date) AS INTEGER) = :m"
        params['m'] = int(month)
    if area:
        query += " AND ward_area = :area"
        params['area'] = area
    query += " GROUP BY ward_area, item_name ORDER BY ward_area, item_name"

    rows = db.session.execute(text(query), params).fetchall()
    return jsonify({'success': True, 'data': [
        {'area': r[0], 'item': r[1], 'sessions': r[2] or 0,
         'amount': float(r[3]) if r[3] else 0.0, 'day_count': r[4]}
        for r in rows
    ]})


@dept_stats_bp.route('/ward/daily/snapshots', methods=['GET'])
def get_ward_daily_snapshots():
    """获取回滚快照列表"""
    ensure_tables()
    area = request.args.get('area', '').strip()
    query = "SELECT id, ward_area, snapshot_time, trigger_type, record_count FROM ward_daily_snapshots WHERE 1=1"
    params = {}
    if area:
        query += " AND ward_area = :area"
        params['area'] = area
    query += " ORDER BY id DESC LIMIT 50"
    rows = db.session.execute(text(query), params).fetchall()
    return jsonify({'success': True, 'data': [
        {'id': r[0], 'ward_area': r[1], 'snapshot_time': str(r[2]),
         'trigger_type': r[3], 'record_count': r[4]}
        for r in rows
    ]})

@dept_stats_bp.route('/ward/daily/rollback', methods=['POST'])
def rollback_ward_daily():
    """从快照恢复病房每日数据"""
    ensure_tables()
    data = request.get_json() or {}
    snapshot_id = int(data.get('snapshot_id') or 0)
    if not snapshot_id:
        return jsonify({'success': False, 'error': '缺少snapshot_id'}), 400
    # 获取快照数据
    row = db.session.execute(text(
        "SELECT ward_area, trigger_type, data_json, record_count FROM ward_daily_snapshots WHERE id = :sid",
        {'sid': snapshot_id}
    )).fetchone()
    if not row:
        return jsonify({'success': False, 'error': '快照不存在'}), 404
    records = json.loads(row[2])
    count = 0
    for rec in records:
        db.session.execute(text("""
            INSERT INTO ward_daily (record_date, ward_area, item_name, session_count, amount)
            VALUES (:date, :area, :name, :sessions, :amt)
            ON CONFLICT(record_date, ward_area, item_name) DO UPDATE SET
                session_count = :sessions, amount = :amt, updated_at = CURRENT_TIMESTAMP
        """), {'date': rec['record_date'], 'area': row[0],
              'name': rec['item_name'],
              'sessions': int(rec.get('session_count', 0)),
              'amt': float(rec.get('amount', 0))})
        count += 1
    db.session.commit()
    # 记录回滚操作日志
    try:
        log_operation('rollback',
            f"病房每日数据回滚: {row[0]} 快照#{snapshot_id} ({row[3]}) 恢复{count}条记录")
    except Exception:
        pass
    return jsonify({'success': True, 'data': {'count': count,
        'area': row[0], 'trigger_type': row[3]}})


@dept_stats_bp.route('/snapshots', methods=['GET'])
def get_data_snapshots():
    """通用快照列表API"""
    ensure_tables()
    module = request.args.get('module', '').strip()
    scope_key = request.args.get('scope_key', '').strip()
    query = ("SELECT id, module, scope_key, table_name, snapshot_time, "
             "trigger_type, record_count FROM data_snapshots WHERE 1=1")
    params = {}
    if module:
        query += " AND module = :mod"
        params['mod'] = module
    if scope_key:
        query += " AND scope_key = :sk"
        params['sk'] = scope_key
    query += " ORDER BY id DESC LIMIT 50"
    rows = db.session.execute(text(query), params).fetchall()
    return jsonify({'success': True, 'data': [
        {'id': r[0], 'module': r[1], 'scope_key': r[2], 'table_name': r[3],
         'snapshot_time': str(r[4]), 'trigger_type': r[5], 'record_count': r[6]}
        for r in rows
    ]})


@dept_stats_bp.route('/rollback', methods=['POST'])
def rollback_data():
    """通用回滚API"""
    ensure_tables()
    data = request.get_json() or {}
    snapshot_id = int(data.get('snapshot_id') or 0)
    if not snapshot_id:
        return jsonify({'success': False, 'error': '缺少snapshot_id'}), 400
    count, err = _do_rollback(snapshot_id)
    if err:
        return jsonify({'success': False, 'error': err}), 404
    # 获取模块信息用于日志
    info = db.session.execute(text(
        "SELECT module, scope_key FROM data_snapshots WHERE id = :sid"
    ), {'sid': snapshot_id}).fetchone()
    try:
        log_operation('rollback',
            f"通用数据回滚: {info[0]} [{info[1]}] 快照#{snapshot_id} 恢复{count}条记录")
    except Exception:
        pass
    return jsonify({'success': True, 'data': {
        'count': count, 'module': info[0], 'scope_key': info[1]
    }})


# ============================================================================
# 四、儿童医院数据
# ============================================================================

@dept_stats_bp.route('/children/monthly', methods=['GET'])
def get_children_monthly():
    ensure_tables()
    year = request.args.get('year', date.today().year, type=int)
    
    rows = db.session.execute(text("""
        SELECT item_name, month, session_count, amount
        FROM children_monthly WHERE year = :year
        ORDER BY month, item_name
    """), {'year': year}).fetchall()
    
    return jsonify({'success': True, 'data': [
        {'item': r[0], 'month': r[1], 'sessions': r[2] or 0, 'amount': r[3] or 0} for r in rows
    ]})


@dept_stats_bp.route('/children/monthly', methods=['POST'])
def save_children_monthly():
    """保存儿童医院月度数据（保存前自动快照）"""
    ensure_tables()
    data = request.get_json()
    year = data.get('year')
    month = data.get('month')
    records = data.get('records', [])
    trigger_type = data.get('trigger_type', 'auto_save')

    if not year or not month or not records:
        return jsonify({'success': False, 'error': '缺少必要参数'}), 400

    # 自动快照
    try:
        _create_snapshot('children_monthly', f'{year}-{month}', 'children_monthly',
            ['item_name', 'year', 'month'],
            "SELECT * FROM children_monthly WHERE year=:y AND month=:m",
            {'y': int(year), 'm': int(month)}, trigger_type=trigger_type)
    except Exception:
        pass

    count = 0
    for r in records:
        name = r.get('item_name', '').strip()
        sessions = r.get('session_count', 0)
        amount = r.get('amount', 0)
        if not name:
            continue
        
        db.session.execute(text("""
            INSERT INTO children_monthly (item_name, year, month, session_count, amount)
            VALUES (:name, :year, :month, :sessions, :amount)
            ON CONFLICT(item_name, year, month) DO UPDATE SET
                session_count = :sessions, amount = :amount, updated_at = CURRENT_TIMESTAMP
        """), {'name': name, 'year': int(year), 'month': int(month), 'sessions': int(sessions), 'amount': float(amount)})
        count += 1
    
    db.session.commit()
    return jsonify({'success': True, 'data': {'count': count}})


@dept_stats_bp.route('/children/daily', methods=['GET'])
def get_children_daily():
    ensure_tables()
    year = request.args.get('year', date.today().year, type=int)
    month = request.args.get('month', type=int)
    
    query = "SELECT record_date, item_name, session_count FROM children_daily WHERE 1=1"
    params = {}
    if year:
        query += " AND strftime('%Y', record_date) = :ys"
        params['ys'] = str(year)
    if month:
        query += " AND CAST(strftime('%m', record_date) AS INTEGER) = :m"
        params['m'] = month
    query += " ORDER BY record_date, item_name"
    
    rows = db.session.execute(text(query), params).fetchall()
    return jsonify({'success': True, 'data': [
        {'date': r[0], 'item': r[1], 'sessions': r[2] or 0} for r in rows
    ]})


@dept_stats_bp.route('/children/daily', methods=['POST'])
def save_children_daily():
    """保存儿童医院每日数据（保存前自动快照）"""
    ensure_tables()
    data = request.get_json()
    record_date = data.get('date')
    records = data.get('records', [])

    if not record_date or not records:
        return jsonify({'success': False, 'error': '缺少必要参数'}), 400

    # 自动快照
    try:
        _create_snapshot('children_daily', str(record_date), 'children_daily',
            ['record_date', 'item_name'],
            "SELECT * FROM children_daily WHERE record_date=:date",
            {'date': record_date})
    except Exception:
        pass

    count = 0
    for r in records:
        name = r.get('item_name', '').strip()
        sessions = r.get('session_count', 0)
        if not name or sessions <= 0:
            continue
        
        db.session.execute(text("""
            INSERT INTO children_daily (record_date, item_name, session_count)
            VALUES (:date, :name, :sessions)
            ON CONFLICT(record_date, item_name) DO UPDATE SET
                session_count = :sessions, updated_at = CURRENT_TIMESTAMP
        """), {'date': record_date, 'name': name, 'sessions': int(sessions)})
        count += 1
    
    db.session.commit()
    return jsonify({'success': True, 'data': {'count': count}})


@dept_stats_bp.route('/children/daily/<int:record_id>', methods=['DELETE'])
def delete_children_daily(record_id):
    ensure_tables()
    db.session.execute(text("DELETE FROM children_daily WHERE id = :id"), {'id': record_id})
    db.session.commit()
    return jsonify({'success': True})


@dept_stats_bp.route('/children/doctor-monthly', methods=['GET'])
def get_children_doctor_monthly():
    ensure_tables()
    year = request.args.get('year', date.today().year, type=int)
    rows = db.session.execute(text("""
        SELECT doctor_name, month, session_count
        FROM children_doctor_monthly WHERE year = :year
        ORDER BY month, doctor_name
    """), {'year': year}).fetchall()
    return jsonify({'success': True, 'data': [
        {'doctor': r[0], 'month': r[1], 'sessions': r[2] or 0} for r in rows
    ]})


@dept_stats_bp.route('/children/doctor-monthly', methods=['POST'])
@log_op('update', '保存儿童医生月度数据')
def save_children_doctor_monthly():
    """保存儿童医生月度数据（保存前自动快照）"""
    ensure_tables()
    data = request.get_json()
    year = data.get('year')
    month = data.get('month')
    records = data.get('records', [])
    trigger_type = data.get('trigger_type', 'auto_save')
    if not year or not month or not records:
        return jsonify({'success': False, 'error': '缺少必要参数'}), 400

    # 自动快照
    try:
        _create_snapshot('children_doctor', f'{year}-{month}', 'children_doctor_monthly',
            ['doctor_name', 'year', 'month'],
            "SELECT * FROM children_doctor_monthly WHERE year=:y AND month=:m",
            {'y': int(year), 'm': int(month)}, trigger_type=trigger_type)
    except Exception:
        pass

    count = 0
    for r in records:
        name = r.get('doctor_name', '').strip()
        sessions = r.get('session_count', 0)
        if not name: continue
        db.session.execute(text("""
            INSERT INTO children_doctor_monthly (doctor_name, year, month, session_count)
            VALUES (:name, :year, :month, :sessions)
            ON CONFLICT(doctor_name, year, month) DO UPDATE SET
                session_count = :sessions, updated_at = CURRENT_TIMESTAMP
        """), {'name': name, 'year': int(year), 'month': int(month), 'sessions': int(sessions)})
        count += 1
    db.session.commit()
    return jsonify({'success': True, 'data': {'count': count}})



@dept_stats_bp.route('/children/monthly/by-item-month', methods=['DELETE'])
def delete_children_monthly_item():
    ensure_tables()
    data = request.get_json() or {}
    item = data.get('item_name', '')
    month = int(data.get('month', 0))
    if not item or not month:
        return jsonify({'success': False, 'error': '参数不完整'}), 400
    db.session.execute(text("DELETE FROM children_monthly WHERE item_name = :item AND month = :month"),
                 {'item': item, 'month': month})
    db.session.commit()
    from app.api.operation_log_bp import log_operation
    log_operation('delete', f'删除儿童月度数据: {item} {month}月')
    return jsonify({'success': True})


@dept_stats_bp.route('/children/monthly/batch-delete', methods=['POST'])
def batch_delete_children_monthly():
    ensure_tables()
    data = request.get_json() or {}
    records = data.get('records', [])
    if not records:
        return jsonify({'success': False, 'error': '请提供要删除的记录'}), 400
    deleted = 0
    for rec in records:
        db.session.execute(text("DELETE FROM children_monthly WHERE item_name = :item AND month = :month"),
                     {'item': rec.get('item_name', ''), 'month': int(rec.get('month', 0))})
        deleted += 1
    db.session.commit()
    from app.api.operation_log_bp import log_operation
    log_operation('delete', f'批量删除儿童月度数据 {deleted}条')
    return jsonify({'success': True, 'deleted': deleted})


@dept_stats_bp.route('/children/doctor-monthly/by-doctor-month', methods=['DELETE'])
def delete_children_doctor_monthly():
    ensure_tables()
    data = request.get_json() or {}
    doctor = data.get('doctor_name', '')
    month = int(data.get('month', 0))
    if not doctor or not month:
        return jsonify({'success': False, 'error': '参数不完整'}), 400
    db.session.execute(text("DELETE FROM children_doctor_monthly WHERE doctor_name = :doc AND month = :month"),
                 {'doc': doctor, 'month': month})
    db.session.commit()
    from app.api.operation_log_bp import log_operation
    log_operation('delete', f'删除儿童医生数据: {doctor} {month}月')
    return jsonify({'success': True})


@dept_stats_bp.route('/children/doctor-monthly/batch-delete', methods=['POST'])
def batch_delete_children_doctor():
    ensure_tables()
    data = request.get_json() or {}
    records = data.get('records', [])
    if not records:
        return jsonify({'success': False, 'error': '请提供要删除的记录'}), 400
    deleted = 0
    for rec in records:
        db.session.execute(text("DELETE FROM children_doctor_monthly WHERE doctor_name = :doc AND month = :month"),
                     {'doc': rec.get('doctor_name', ''), 'month': int(rec.get('month', 0))})
        deleted += 1
    db.session.commit()
    from app.api.operation_log_bp import log_operation
    log_operation('delete', f'批量删除儿童医生数据 {deleted}条')
    return jsonify({'success': True, 'deleted': deleted})


@dept_stats_bp.route('/children/clear', methods=['POST'])
def clear_children():
    ensure_tables()
    data = request.get_json() or {}
    scope = data.get('scope', 'all')
    if scope == 'daily':
        db.session.execute(text("DELETE FROM children_daily"))
    elif scope == 'monthly':
        db.session.execute(text("DELETE FROM children_monthly"))
    else:
        db.session.execute(text("DELETE FROM children_daily"))
        db.session.execute(text("DELETE FROM children_monthly"))
    db.session.commit()
    return jsonify({'success': True})


# ============================================================================
# 五、图表数据（聚合查询）
# ============================================================================

def _row_to_dict(r):
    """SQLAlchemy Row -> dict"""
    if hasattr(r, '_asdict'):
        return dict(r._asdict())
    elif hasattr(r, 'keys'):
        return dict(zip(r.keys(), tuple(r)))
    else:
        return dict(r)

def _rows_to_dicts(rows):
    """List of SQLAlchemy Rows -> list of dicts"""
    return [_row_to_dict(r) for r in rows]


@dept_stats_bp.route('/charts', methods=['GET'])
def get_charts():
    ensure_tables()
    year = request.args.get('year', date.today().year, type=int)
    ward_daily_area = request.args.get('ward_daily_area', '').strip()
    ward_daily_month = request.args.get('month', '', type=str).strip()

    result = {}

    # 门诊月度合计
    result['outpatient_monthly'] = _rows_to_dicts(db.session.execute(text("""
        SELECT month, SUM(session_count) as sessions, SUM(amount) as amount
        FROM outpatient_monthly WHERE year = :y GROUP BY month ORDER BY month
    """), {'y': year}).fetchall())
    
    # 门诊项目排名
    result['outpatient_ranking'] = _rows_to_dicts(db.session.execute(text("""
        SELECT item_name as item, SUM(session_count) as sessions, SUM(amount) as amount
        FROM outpatient_monthly WHERE year = :y
        GROUP BY item_name ORDER BY sessions DESC LIMIT 15
    """), {'y': year}).fetchall())
    
    # 病房月度合计
    result['ward_monthly'] = _rows_to_dicts(db.session.execute(text("""
        SELECT month, SUM(session_count) as sessions, SUM(amount) as amount FROM ward_monthly
        WHERE year = :y GROUP BY month ORDER BY month
    """), {'y': year}).fetchall())
    
    # 病房项目排名
    result['ward_ranking'] = _rows_to_dicts(db.session.execute(text("""
        SELECT ward_area as area, item_name as item, SUM(session_count) as sessions, SUM(amount) as amount
        FROM ward_monthly WHERE year = :y
        GROUP BY ward_area, item_name ORDER BY sessions DESC LIMIT 15
    """), {'y': year}).fetchall())
    
    # 病房按病区合计
    result['ward_area_total'] = _rows_to_dicts(db.session.execute(text("""
        SELECT ward_area as area, SUM(session_count) as sessions, SUM(amount) as amount
        FROM ward_monthly WHERE year = :y
        GROUP BY ward_area ORDER BY sessions DESC
    """), {'y': year}).fetchall())
    
    # 儿童医院月度合计
    result['children_monthly'] = _rows_to_dicts(db.session.execute(text("""
        SELECT month, SUM(session_count) as sessions FROM children_monthly
        WHERE year = :y GROUP BY month ORDER BY month
    """), {'y': year}).fetchall())
    
    # 儿童医院项目排名
    result['children_ranking'] = _rows_to_dicts(db.session.execute(text("""
        SELECT item_name as item, SUM(session_count) as sessions
        FROM children_monthly WHERE year = :y
        GROUP BY item_name ORDER BY sessions DESC LIMIT 15
    """), {'y': year}).fetchall())
    
    # 医生开单月度合计
    result['doctor_monthly'] = _rows_to_dicts(db.session.execute(text("""
        SELECT month, SUM(session_count) as sessions FROM children_doctor_monthly
        WHERE year = :y GROUP BY month ORDER BY month
    """), {'y': year}).fetchall())
    
    # 医生开单排名
    result['doctor_ranking'] = _rows_to_dicts(db.session.execute(text("""
        SELECT doctor_name as doctor, SUM(session_count) as sessions
        FROM children_doctor_monthly WHERE year = :y
        GROUP BY doctor_name ORDER BY sessions DESC LIMIT 15
    """), {'y': year}).fetchall())

    # 儿童医院月度合计(补全金额)
    result['children_monthly'] = _rows_to_dicts(db.session.execute(text("""
        SELECT month, SUM(session_count) as sessions, SUM(amount) as amount FROM children_monthly
        WHERE year = :y GROUP BY month ORDER BY month
    """), {'y': year}).fetchall())

    # 儿童项目排名(补全金额)
    result['children_ranking'] = _rows_to_dicts(db.session.execute(text("""
        SELECT item_name as item, SUM(session_count) as sessions, SUM(amount) as amount
        FROM children_monthly WHERE year = :y
        GROUP BY item_name ORDER BY sessions DESC LIMIT 15
    """), {'y': year}).fetchall())

    # 病房每日趋势（支持按科室+月份过滤）
    trend_params = {'y': str(year)}
    trend_where = "WHERE strftime('%Y', record_date) = :y"
    if ward_daily_area:
        trend_where += " AND ward_area = :wda"
        trend_params['wda'] = ward_daily_area
    if ward_daily_month:
        trend_where += " AND strftime('%m', record_date) = :wdm"
        trend_params['wdm'] = ward_daily_month.zfill(2)
    result['ward_daily_trend'] = _rows_to_dicts(db.session.execute(text(f"""
        SELECT record_date as date, SUM(session_count) as sessions, SUM(amount) as amount
        FROM ward_daily {trend_where}
        GROUP BY record_date ORDER BY record_date
    """), trend_params).fetchall())

    # 病房每日项目排名（支持按科室+月份过滤）
    rank_params = {'y': str(year)}
    rank_where = "WHERE strftime('%Y', record_date) = :y"
    if ward_daily_area:
        rank_where += " AND ward_area = :wda"
        rank_params['wda'] = ward_daily_area
    if ward_daily_month:
        rank_where += " AND strftime('%m', record_date) = :wdm"
        rank_params['wdm'] = ward_daily_month.zfill(2)
    result['ward_daily_item_ranking'] = _rows_to_dicts(db.session.execute(text(f"""
        SELECT item_name as item, SUM(session_count) as sessions, SUM(amount) as amount
        FROM ward_daily {rank_where}
        GROUP BY item_name ORDER BY sessions DESC
    """), rank_params).fetchall())

    return jsonify({'success': True, 'data': result})


# ============================================================================
# 六、转介统计
# ============================================================================

@referral_bp.route('/doctors', methods=['GET'])
def get_referral_doctors():
    ensure_tables()
    dept = request.args.get('department', '')
    query = "SELECT id, department, doctor_name FROM referral_doctors WHERE is_active = 1"
    params = {}
    if dept:
        query += " AND department = :dept"
        params['dept'] = dept
    query += " ORDER BY department, doctor_name"
    
    rows = db.session.execute(text(query), params).fetchall()
    return jsonify({'success': True, 'data': [
        {'id': r[0], 'department': r[1], 'name': r[2]} for r in rows
    ]})


@referral_bp.route('/doctors', methods=['POST'])
def add_referral_doctor():
    ensure_tables()
    data = request.get_json()
    dept = data.get('department', '').strip()
    name = data.get('name', '').strip()
    if not dept or not name:
        return jsonify({'success': False, 'error': '科室和医生名不能为空'}), 400
    
    db.session.execute(text("""
        INSERT OR IGNORE INTO referral_doctors (department, doctor_name) VALUES (:dept, :name)
    """), {'dept': dept, 'name': name})
    db.session.commit()
    return jsonify({'success': True})


@referral_bp.route('/doctors/<int:doc_id>', methods=['DELETE'])
def delete_referral_doctor(doc_id):
    ensure_tables()
    # 先获取医生信息，用于清理关联数据
    row = db.session.execute(text("SELECT department, doctor_name FROM referral_doctors WHERE id = :id"), {'id': doc_id}).fetchone()
    if row:
        dept, name = row[0], row[1]
        # 清理该医生的月度数据
        db.session.execute(text("DELETE FROM referral_monthly WHERE department = :dept AND doctor_name = :doc"), {'dept': dept, 'doc': name})
    db.session.execute(text("DELETE FROM referral_doctors WHERE id = :id"), {'id': doc_id})
    db.session.commit()
    return jsonify({'success': True})


@referral_bp.route('/doctors/batch', methods=['POST'])
def batch_add_referral_doctors():
    ensure_tables()
    data = request.get_json()
    dept = data.get('department', '').strip()
    names = data.get('names', [])
    if not dept:
        return jsonify({'success': False, 'error': '缺少科室'}), 400
    
    count = 0
    for name in names:
        name = name.strip()
        if not name:
            continue
        db.session.execute(text("""
            INSERT OR IGNORE INTO referral_doctors (department, doctor_name) VALUES (:dept, :name)
        """), {'dept': dept, 'name': name})
        count += 1
    db.session.commit()
    return jsonify({'success': True, 'data': {'count': count}})


@referral_bp.route('/metrics', methods=['GET'])
def get_referral_metrics():
    ensure_tables()
    dept = request.args.get('department', '')
    query = "SELECT id, department, metric_name FROM referral_metrics WHERE is_active = 1"
    params = {}
    if dept:
        query += " AND department = :dept"
        params['dept'] = dept
    query += " ORDER BY metric_name"
    
    rows = db.session.execute(text(query), params).fetchall()
    return jsonify({'success': True, 'data': [
        {'id': r[0], 'department': r[1], 'name': r[2]} for r in rows
    ]})


@referral_bp.route('/metrics', methods=['POST'])
def add_referral_metric():
    ensure_tables()
    data = request.get_json()
    dept = data.get('department', '').strip()
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'error': '指标名不能为空'}), 400
    
    db.session.execute(text("""
        INSERT OR IGNORE INTO referral_metrics (department, metric_name) VALUES (:dept, :name)
    """), {'dept': dept, 'name': name})
    db.session.commit()
    return jsonify({'success': True})


@referral_bp.route('/metrics/batch', methods=['POST'])
def batch_add_referral_metrics():
    ensure_tables()
    data = request.get_json()
    dept = data.get('department', '').strip()
    names = data.get('names', [])
    
    count = 0
    for name in names:
        name = name.strip()
        if not name:
            continue
        db.session.execute(text("""
            INSERT OR IGNORE INTO referral_metrics (department, metric_name) VALUES (:dept, :name)
        """), {'dept': dept, 'name': name})
        count += 1
    db.session.commit()
    return jsonify({'success': True, 'data': {'count': count}})


@referral_bp.route('/departments', methods=['GET'])
def get_departments():
    ensure_tables()
    # 优先从referral_doctors取，再从referral_monthly补充（覆盖import-json写入的场景）
    rows = db.session.execute(text("""
        SELECT DISTINCT department FROM referral_doctors WHERE is_active = 1
        UNION
        SELECT DISTINCT department FROM referral_monthly WHERE department != ''
        ORDER BY department
    """)).fetchall()
    return jsonify({'success': True, 'data': [r[0] for r in rows]})


# --- 转介数据 CRUD ---

@referral_bp.route('/data', methods=['GET'])
def get_referral_data():
    ensure_tables()
    year = request.args.get('year', date.today().year, type=int)
    dept = request.args.get('department', '')
    
    query = "SELECT department, doctor_name, month, metric_name, metric_count FROM referral_monthly WHERE year = :y"
    params = {'y': year}
    if dept:
        query += " AND department = :dept"
        params['dept'] = dept
    query += " ORDER BY department, doctor_name, month"
    
    rows = db.session.execute(text(query), params).fetchall()
    return jsonify({'success': True, 'data': [
        {'department': r[0], 'doctor': r[1], 'month': r[2], 'metric': r[3], 'count': r[4] or 0}
        for r in rows
    ]})


@referral_bp.route('/data', methods=['POST'])
@log_op('update', '保存转介数据')
def save_referral_data():
    """保存转介月度数据（保存前自动快照）"""
    ensure_tables()
    data = request.get_json()
    year = data.get('year')
    month = data.get('month')
    records = data.get('records', [])
    trigger_type = data.get('trigger_type', 'auto_save')

    if not year or not records:
        return jsonify({'success': False, 'error': '缺少必要参数'}), 400

    # 自动快照
    try:
        y=int(year)
        if int(month or 0) > 0:
            _create_snapshot('referral', f'{y}-{month}', 'referral_monthly',
                ['department', 'doctor_name', 'year', 'month', 'metric_name'],
                "SELECT * FROM referral_monthly WHERE year=:y AND month=:m",
                {'y': y, 'm': int(month)}, trigger_type=trigger_type)
        else:
            _create_snapshot('referral', str(y), 'referral_monthly',
                ['department', 'doctor_name', 'year', 'month', 'metric_name'],
                "SELECT * FROM referral_monthly WHERE year=:y",
                {'y': y}, trigger_type=trigger_type)
    except Exception:
        pass

    count = 0
    for r in records:
        dept = r.get('department', '').strip()
        doctor = r.get('doctor_name', '').strip()
        metric = r.get('metric_name', '').strip()
        mcount = r.get('metric_count', 0)
        if not dept or not doctor or not metric:
            continue
        
        db.session.execute(text("""
            INSERT INTO referral_monthly (department, doctor_name, year, month, metric_name, metric_count)
            VALUES (:dept, :doctor, :year, :month, :metric, :count)
            ON CONFLICT(department, doctor_name, year, month, metric_name) DO UPDATE SET
                metric_count = :count, updated_at = CURRENT_TIMESTAMP
        """), {'dept': dept, 'doctor': doctor, 'year': int(year), 'month': int(month),
              'metric': metric, 'count': int(mcount)})
        count += 1
    
    db.session.commit()
    return jsonify({'success': True, 'data': {'count': count}})


@referral_bp.route('/data/<int:record_id>', methods=['DELETE'])
def delete_referral_data(record_id):
    ensure_tables()
    db.session.execute(text("DELETE FROM referral_monthly WHERE id = :id"), {'id': record_id})
    db.session.commit()
    return jsonify({'success': True})


@referral_bp.route('/data/delete-by-key', methods=['POST'])
@log_op('delete', '删除转介记录')
def delete_referral_by_key():
    """按科室+医生删除转介月度记录"""
    ensure_tables()
    data = request.get_json() or {}
    year = int(data.get('year') or 0)
    department = data.get('department', '')
    doctor_name = data.get('doctor_name', '')
    if not year or not department:
        return jsonify({'success': False, 'error': '缺少参数'}), 400
    params = {'y': int(year), 'dept': department}
    sql = "DELETE FROM referral_monthly WHERE year = :y AND department = :dept"
    if doctor_name:
        sql += " AND doctor_name = :doc"
        params['doc'] = doctor_name
    result = db.session.execute(text(sql), params)
    db.session.commit()
    return jsonify({'success': True, 'data': {'deleted_count': result.rowcount}})


@referral_bp.route('/data/clear', methods=['POST'])
def clear_referral_data():
    ensure_tables()
    data = request.get_json() or {}
    year = data.get('year')
    month = data.get('month')
    if year and month:
        db.session.execute(text("DELETE FROM referral_monthly WHERE year = :y AND month = :m"),
                         {'y': int(year), 'm': int(month)})
    else:
        db.session.execute(text("DELETE FROM referral_monthly"))
    db.session.commit()
    return jsonify({'success': True})


@referral_bp.route('/charts', methods=['GET'])
def get_referral_charts():
    ensure_tables()
    year = request.args.get('year', date.today().year, type=int)
    dept = request.args.get('department', '')
    
    where = "WHERE year = :y"
    params = {'y': year}
    if dept:
        where += " AND department = :dept"
        params['dept'] = dept
    
    result = {}
    
    result['doctor_ranking'] = _rows_to_dicts(db.session.execute(text(f"""
        SELECT doctor_name as doctor, department, SUM(metric_count) as total
        FROM referral_monthly {where}
        GROUP BY doctor_name, department ORDER BY total DESC
    """), params).fetchall())
    
    result['dept_monthly'] = _rows_to_dicts(db.session.execute(text(f"""
        SELECT department, month, SUM(metric_count) as total
        FROM referral_monthly {where}
        GROUP BY department, month ORDER BY department, month
    """), params).fetchall())
    
    result['dept_yearly'] = _rows_to_dicts(db.session.execute(text(f"""
        SELECT department, SUM(metric_count) as total
        FROM referral_monthly {where}
        GROUP BY department ORDER BY total DESC
    """), params).fetchall())
    
    result['doctor_monthly'] = _rows_to_dicts(db.session.execute(text(f"""
        SELECT doctor_name as doctor, department, month, SUM(metric_count) as total
        FROM referral_monthly {where}
        GROUP BY doctor_name, department, month ORDER BY doctor_name, department, month
    """), params).fetchall())
    
    return jsonify({'success': True, 'data': result})


@referral_bp.route('/template', methods=['GET'])
def download_referral_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '转介数据'
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 10
    header_fill = openpyxl.styles.PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    for ci, h in enumerate(['科室', '医生姓名', '人次'], start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
    thin_border = openpyxl.styles.Side(style='thin')
    border = openpyxl.styles.Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    for r in range(1, 6):
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = border
    ws2 = wb.create_sheet('填写说明')
    ws2.column_dimensions['A'].width = 60
    for i, line in enumerate([
        '【填写说明】', '',
        'A列填写科室，B列填写医生姓名，C列填写人次',
        '科室只在第一次出现时填写，后续留空会自动继承上一个科室',
        '',
        '用法一：仅导入医生',
        '  A列填科室，B列填医生姓名，C列留空，在录入框中导入即可',
        '',
        '用法二：导入医生+月度数据',
        '  在录入框中选好年份和月份，然后点「导入数据Excel」',
        '', '提示：系统会自动去重。月度数据如已存在会自动覆盖。'
    ], start=1):
        ws2.cell(row=i, column=1, value=line)
        if i == 1: ws2.cell(row=i, column=1).font = openpyxl.styles.Font(bold=True, size=14)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='转介数据导入模板.xlsx')


@referral_bp.route('/monthly', methods=['GET'])
def get_referral_monthly():
    """获取转介月度数据"""
    ensure_tables()
    year = request.args.get('year', '', type=str)
    month = request.args.get('month', type=int)
    dept = request.args.get('department', '')
    
    sql = "SELECT MIN(id) as id, year, month, department, doctor_name, metric_name, SUM(metric_count) AS total_sessions, MAX(updated_at) AS updated_at FROM referral_monthly WHERE 1=1"
    params = {}
    if year:
        sql += " AND year = :y"
        params['y'] = year
    if month is not None:
        sql += " AND month = :m"
        params['m'] = month
    if dept:
        sql += " AND department = :d"
        params['d'] = dept
    sql += " GROUP BY department, doctor_name, year, month"
    
    rows = db.session.execute(text(sql), params).fetchall()
    return jsonify({'success': True, 'data': _rows_to_dicts(rows)})


@referral_bp.route('/monthly/import-excel', methods=['POST'])
@log_op('import', '转介导入Excel')
def import_referral_monthly_excel():
    ensure_tables()
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '请上传文件'}), 400
    year = request.form.get('year', '').strip()
    months_str = request.form.get('months', '[]')
    try: months = json.loads(months_str)
    except: months = []
    print(f'[DEBUG-IMPORT] parsed months={months} type(months)={type(months)}')
    file = request.files['file']
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        count = 0; monthly_count = 0
        skip = ('科室', '医生姓名', '医生', '姓名', 'name', '人次', '数量')
        for ws in wb.worksheets:
            current_dept = None
            for row in ws.iter_rows(values_only=True):
                if not row: continue
                col_a = str(row[0]).strip() if row[0] else ''
                col_b = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                if col_a in skip or col_b in skip: continue
                # 科室继承（A列有值则更新，空则继承上一个科室）
                if col_a and col_b:
                    current_dept = col_a
                    name = col_b
                elif col_a and not col_b:
                    # A列有值B列空 = 科室标题行，更新继承后跳过
                    current_dept = col_a
                    continue
                else:
                    name = col_b
                if not name or not current_dept: continue
                try: float(name); continue
                except: pass
                sessions = 0
                if len(row) > 2 and row[2] is not None and str(row[2]).strip() != '':
                    try: sessions = int(float(str(row[2]).strip()))
                    except: pass
                # 注册医生（科室信息会随医生一起存储）
                db.session.execute(text("INSERT OR IGNORE INTO referral_doctors (department, doctor_name) VALUES (:dept, :name)"), {'dept': current_dept, 'name': name})
                count += 1
                # 写入月度数据（允许人次数为0）
                if months and year:
                    for mon in months:
                        if mon < 1 or mon > 12: continue
                        db.session.execute(text("""
                            INSERT INTO referral_monthly (department, doctor_name, year, month, metric_name, metric_count)
                            VALUES (:dept, :name, :year, :month, :metric, :count)
                            ON CONFLICT(department, doctor_name, year, month, metric_name) DO UPDATE SET
                                metric_count = :count, updated_at = CURRENT_TIMESTAMP
                        """), {'dept': current_dept, 'name': name, 'year': int(year), 'month': int(mon), 'metric': '转介', 'count': sessions})
                        monthly_count += 1
        db.session.commit()
        return jsonify({'success': True, 'data': {'count': count, 'monthly_count': monthly_count}})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@referral_bp.route('/monthly/import-json', methods=['POST'])
@log_op('import', '转介导入JSON')
def import_referral_monthly_json():
    ensure_tables()
    try:
        # Support both FormData (frontend sends records array as 'data' field)
        # and raw JSON API ({records:[...], year:'...', month:N})
        raw_data = request.form.get('data')
        if raw_data:
            parsed = json.loads(raw_data)
            # Frontend sends: fd.append('data', JSON.stringify(records))
            # So parsed is the records list directly
            if isinstance(parsed, list):
                records = parsed
                year = str(request.form.get('year', date.today().year))
                month = int(request.form.get('month') or 1)
            else:
                records = parsed.get('records', [])
                year = str(parsed.get('year', request.form.get('year', date.today().year)))
                month = int(parsed.get('month') or request.form.get('month') or 1)
        else:
            data = request.get_json(force=True) or {}
            records = data.get('records', [])
            year = str(data.get('year', date.today().year))
            month = int(data.get('month') or 1)
        # 参数校验
        if not year or not year.isdigit() or len(year) != 4:
            return jsonify({'success': False, 'error': '年份参数无效'}), 400
        if not month or month < 1 or month > 12:
            return jsonify({'success': False, 'error': '月份参数无效（1-12）'}), 400
        count = 0
        for r in records:
            dept = (r.get('department') or '').strip()
            doctor = (r.get('doctor_name') or '').strip()
            sessions = int(r.get('sessions') or 0)
            if not doctor:
                continue
            # 同步注册医生到referral_doctors表（用于科室筛选下拉框）
            if dept:
                db.session.execute(text(
                    "INSERT OR IGNORE INTO referral_doctors (department, doctor_name) VALUES (:dept, :name)"
                ), {'dept': dept, 'name': doctor})
            exists = db.session.execute(text(
                "SELECT id FROM referral_monthly WHERE year=:y AND month=:m AND department=:d AND doctor_name=:doc"
            ), {'y': year, 'm': month, 'd': dept, 'doc': doctor}).fetchone()
            if exists:
                db.session.execute(text(
                    "UPDATE referral_monthly SET metric_name='转介',metric_count=:c WHERE year=:y AND month=:m AND department=:d AND doctor_name=:doc"
                ), {'c': sessions, 'y': year, 'm': month, 'd': dept, 'doc': doctor})
            else:
                db.session.execute(text(
                    "INSERT INTO referral_monthly (year,month,department,doctor_name,metric_name,metric_count) VALUES(:y,:m,:d,:doc,'转介',:c)"
                ), {'y': year, 'm': month, 'd': dept, 'doc': doctor, 'c': sessions})
            count += 1
        db.session.commit()
        return jsonify({'success': True, 'data': {'count': count}})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500



@dept_stats_bp.route('/init', methods=['GET'])
def batch_init():
    """批量初始化接口：一次请求返回科室统计模块需要的所有基础数据"""
    ensure_tables()
    year = request.args.get('year', date.today().year, type=int)
    month = request.args.get('month', type=int)
    result = {}

    # 1. 门诊项目列表
    items_rows = db.session.execute(text("""
        SELECT item_name, sort_order FROM dept_items
        WHERE item_type='outpatient' ORDER BY sort_order, item_name
    """)).fetchall()
    result['outpatient_items'] = [{'name': r[0], 'sort': r[1]} for r in items_rows]

    # 2. 病区列表
    areas_rows = db.session.execute(text("""
        SELECT DISTINCT ward_area FROM ward_daily WHERE ward_area IS NOT NULL AND ward_area != ''
        UNION
        SELECT DISTINCT ward_area FROM ward_item_prices WHERE ward_area IS NOT NULL AND ward_area != ''
        ORDER BY ward_area
    """)).fetchall()
    result['ward_areas'] = [r[0] for r in areas_rows]
    if not result['ward_areas']:
        result['ward_areas'] = ['产科', '妇科', '外科', '新生儿科', '产康中心']

    # 3. 转介科室
    ref_rows = db.session.execute(text("""
        SELECT DISTINCT department FROM referral_monthly ORDER BY department
    """)).fetchall()
    result['referral_departments'] = [r[0] for r in ref_rows]

    # 4-5. 当前月数据（如果指定了月份）
    if month:
        out_rows = db.session.execute(text("""
            SELECT item_name, session_count, amount FROM outpatient_monthly
            WHERE year=:year AND month=:month ORDER BY item_name
        """), {'year': year, 'month': month}).fetchall()
        result['outpatient_data'] = [
            {'item': r[0], 'sessions': r[1] or 0, 'amount': r[2] or 0} for r in out_rows
        ]
        summary_rows = db.session.execute(text("""
            SELECT ward_area, item_name,
                   COALESCE(SUM(session_count),0) as sessions,
                   COALESCE(SUM(amount),0) as amount
            FROM ward_daily
            WHERE strftime('%Y', record_date)=:year_str AND strftime('%m', record_date)=:month_str
            GROUP BY ward_area, item_name
            ORDER BY ward_area, item_name
        """), {'year_str': str(year), 'month_str': f'{month:02d}'}).fetchall()
        result['ward_summary'] = [
            {'area': r[0], 'item': r[1], 'sessions': r[2], 'amount': r[3]} for r in summary_rows
        ]

    return jsonify({'success': True, 'data': result})
